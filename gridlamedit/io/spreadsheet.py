"""Spreadsheet import pipeline and UI binding for GridLamEdit."""

from __future__ import annotations

import logging
import math
import re
import unicodedata
import zipfile
from collections import Counter, OrderedDict
from contextlib import contextmanager
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Dict, Iterable, List, Optional, Protocol, Sequence

import pandas as pd
from PySide6.QtCore import (
    QAbstractItemModel,
    QAbstractTableModel,
    QModelIndex,
    QPoint,
    QRect,
    Qt,
)
from PySide6.QtGui import QColor, QPalette, QUndoCommand, QUndoStack, QIcon
from PySide6.QtWidgets import (
    QAbstractItemView,
    QComboBox,
    QHeaderView,
    QStyle,
    QStyleOptionHeader,
    QStyleOptionButton,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QSizePolicy,
    QTableView,
    QWidget,
)

_OPEN_EXCEL_EXCEPTIONS: tuple[type[BaseException], ...] = (zipfile.BadZipFile,)
try:  # pragma: no cover - dependente de openpyxl
    from openpyxl.utils.exceptions import InvalidFileException
except Exception:  # pragma: no cover - fallback quando openpyxl nAo disponAvel
    InvalidFileException = None  # type: ignore[assignment]
else:  # pragma: no cover - depende de openpyxl
    _OPEN_EXCEL_EXCEPTIONS = _OPEN_EXCEL_EXCEPTIONS + (InvalidFileException,)  # type: ignore[assignment]

try:  # pragma: no cover - dependente de xlrd
    import xlrd  # type: ignore
except Exception:  # pragma: no cover - fallback quando xlrd nAo disponAvel
    xlrd = None  # type: ignore[assignment]

logger = logging.getLogger(__name__)

ORIENTATION_MIN = -100.0
ORIENTATION_MAX = 100.0
_DEGREE_TOKENS = ("\N{DEGREE SIGN}", "\u00ba")
_ORIENTATION_TEXT_PATTERN = re.compile(r"^[+-]?\d+(?:[.,]\d+)?$")
ORIENTATION_HIGHLIGHT_COLORS: dict[float, QColor] = {
    45.0: QColor(193, 174, 255),  # Lil├ís
    90.0: QColor(160, 196, 255),  # Azul
    -45.0: QColor(176, 230, 176),  # Verde claro
    0.0: QColor(230, 230, 230),  # Cinza claro
}
DEFAULT_ORIENTATION_HIGHLIGHT = QColor(255, 236, 200)

LAMINATE_ALIASES = ("Laminate", "Laminate Name", "Laminado", "Nome")
COLOR_ALIASES = ("Color", "Colour", "Cor", "ColorIdx", "Color Index")
TYPE_ALIASES = ("Type", "Tipo")
TAG_ALIASES = ("Tag",)
MATERIAL_ALIASES = ("Material",)
ORIENTATION_ALIASES = ("Orientation", "Orientacao", "Orientacao", "Angle", "Angulo", "Angulo")
ACTIVE_ALIASES = ("Active", "Ativo", "Status")
SYMMETRY_ALIASES = ("Symmetry", "Simetria")
INDEX_ALIASES = ("Index", "#", "Idx", "Ordem", "SequAancia", "Sequencia")

CELL_MAPPING_ALIASES = ("Cell", "Cells", "Celula", "CAlula", "C")
CELL_ID_PATTERN = re.compile(r"^C\d+$", re.IGNORECASE)
NO_LAMINATE_LABEL = "(sem laminado)"
PLY_TYPE_OPTIONS: tuple[str, str] = ("Considerar", "N\u00e3o Considerar")
DEFAULT_PLY_TYPE = PLY_TYPE_OPTIONS[0]
DEFAULT_ROSETTE_LABEL = "Rosette.1"
DEFAULT_COLOR_INDEX = 1
MIN_COLOR_INDEX = 1
MAX_COLOR_INDEX = 150
_HEX_COLOR_PATTERN = re.compile(r"#?[0-9a-fA-F]{6}")
ORIENTATION_SYMMETRY_ROLE = Qt.UserRole + 50  # Custom role to signal symmetric pairing.


def _normalize_ply_type_token(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    normalized = unicodedata.normalize("NFKD", text)
    ascii_only = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    ascii_only = ascii_only.lower()
    return re.sub(r"[^a-z0-9]+", "", ascii_only)


# Mantem compatibilidade com valores antigos (ex.: "Structural Ply").
_PLY_TYPE_CANONICAL_MAP: dict[str, str] = {
    _normalize_ply_type_token("Structural Ply"): PLY_TYPE_OPTIONS[0],
    _normalize_ply_type_token("Structural"): PLY_TYPE_OPTIONS[0],
    _normalize_ply_type_token("Considerar"): PLY_TYPE_OPTIONS[0],
    _normalize_ply_type_token("Nonstructural Ply"): PLY_TYPE_OPTIONS[1],
    _normalize_ply_type_token("Non structural Ply"): PLY_TYPE_OPTIONS[1],
    _normalize_ply_type_token("Nonstructural"): PLY_TYPE_OPTIONS[1],
    _normalize_ply_type_token("Nao Considerar"): PLY_TYPE_OPTIONS[1],
    _normalize_ply_type_token("N\u00e3o Considerar"): PLY_TYPE_OPTIONS[1],
}


def normalize_ply_type_label(value: object) -> str:
    token = _normalize_ply_type_token(value)
    if not token:
        return DEFAULT_PLY_TYPE
    return _PLY_TYPE_CANONICAL_MAP.get(token, DEFAULT_PLY_TYPE)


def is_structural_ply_label(value: object) -> bool:
    return normalize_ply_type_label(value) == PLY_TYPE_OPTIONS[0]


def is_known_ply_type_value(value: object) -> bool:
    token = _normalize_ply_type_token(value)
    return bool(token) and token in _PLY_TYPE_CANONICAL_MAP


def ply_type_signature_token(value: object) -> str:
    token = _normalize_ply_type_token(normalize_ply_type_label(value))
    return token or _normalize_ply_type_token(DEFAULT_PLY_TYPE)


class WordWrapHeader(QHeaderView):
    def __init__(
        self,
        orientation: Qt.Orientation,
        parent: Optional[QWidget] = None,
        *,
        checkbox_section: Optional[int] = None,
    ) -> None:
        super().__init__(orientation, parent)
        self.setDefaultAlignment(Qt.AlignCenter)
        self.setSectionsClickable(True)
        self._checkbox_section: Optional[int] = checkbox_section
        self._connected_model: Optional[QAbstractItemModel] = None
        self._signal_handlers = {
            "headerDataChanged": self._on_header_changed,
            "modelReset": self._on_model_reset,
            "layoutChanged": self._on_model_reset,
        }

    def set_checkbox_section(self, section: Optional[int]) -> None:
        previous = self._checkbox_section
        self._checkbox_section = section
        if previous is not None:
            self.updateSection(previous)
        if section is not None:
            self.updateSection(section)
        else:
            self.viewport().update()

    def checkbox_section(self) -> Optional[int]:
        return self._checkbox_section

    def setModel(self, model: Optional[QAbstractItemModel]) -> None:
        previous = self.model()
        if previous is not None:
            self._disconnect_model_signals(previous)
        super().setModel(model)
        self._connect_model_signals(model)
        self._on_model_reset()

    def _connect_model_signals(self, model: Optional[QAbstractItemModel]) -> None:
        if model is None:
            self._connected_model = None
            return
        if self._connected_model is model:
            return
        self._connected_model = model
        for signal_name, handler in self._signal_handlers.items():
            try:
                getattr(model, signal_name).connect(handler)
            except AttributeError:
                continue

    def _disconnect_model_signals(self, model: Optional[QAbstractItemModel]) -> None:
        if model is None:
            return
        for signal_name, handler in self._signal_handlers.items():
            try:
                getattr(model, signal_name).disconnect(handler)
            except (AttributeError, TypeError, RuntimeError):
                continue

    def _on_header_changed(self, orientation: Qt.Orientation, first: int, last: int) -> None:
        if orientation != Qt.Horizontal:
            return
        for section in range(first, last + 1):
            self.updateSection(section)
        self.updateGeometry()
        self.viewport().update()

    def _on_model_reset(self, *args) -> None:
        if self.model() is None:
            return
        self.updateGeometry()
        self.viewport().update()

    def paintSection(self, painter, rect, logicalIndex):  # noqa: N802
        if not rect.isValid():
            return
        painter.save()
        option = QStyleOptionHeader()
        self.initStyleOption(option)
        option.rect = rect
        option.section = logicalIndex
        option.text = ""
        self.style().drawControl(QStyle.CE_Header, option, painter, self)

        model = self.model()
        header_text = ""
        decoration = None
        if model is not None:
            text = model.headerData(logicalIndex, self.orientation(), Qt.DisplayRole)
            if text:
                header_text = str(text)
            decoration = model.headerData(logicalIndex, self.orientation(), Qt.DecorationRole)

        decoration_pixmap = None
        if isinstance(decoration, QIcon):
            decoration_pixmap = decoration.pixmap(18, 18)
        elif hasattr(decoration, "isNull"):
            try:
                if not decoration.isNull():
                    decoration_pixmap = decoration
            except Exception:
                decoration_pixmap = None

        painter.setPen(option.palette.color(QPalette.ButtonText))
        icon_bottom = rect.top()
        if decoration_pixmap is not None and not decoration_pixmap.isNull():
            icon_rect = QRect(decoration_pixmap.rect())
            icon_rect.moveCenter(
                QPoint(rect.center().x(), rect.top() + icon_rect.height() // 2 + 2)
            )
            painter.drawPixmap(icon_rect, decoration_pixmap)
            icon_bottom = icon_rect.bottom()

        if (
            self.orientation() == Qt.Horizontal
            and self._checkbox_section is not None
            and logicalIndex == self._checkbox_section
        ):
            checkbox_rect = self._calculate_checkbox_rect(rect)

            option_button = QStyleOptionButton()
            option_button.state = QStyle.State_Enabled

            check_state = Qt.Unchecked
            if model is not None and hasattr(model, "all_checked") and hasattr(model, "any_checked"):
                try:
                    if model.all_checked():
                        check_state = Qt.Checked
                    elif model.any_checked():
                        check_state = Qt.PartiallyChecked
                except Exception:  # pragma: no cover - fallback defensivo
                    check_state = Qt.Unchecked

            if check_state == Qt.Checked:
                option_button.state |= QStyle.State_On
            elif check_state == Qt.PartiallyChecked:
                option_button.state |= QStyle.State_NoChange
            else:
                option_button.state |= QStyle.State_Off
            option_button.rect = checkbox_rect
            self.style().drawControl(QStyle.CE_CheckBox, option_button, painter, self)

            text_rect = QRect(
                rect.left() + 4,
                max(checkbox_rect.bottom(), icon_bottom) + 4,
                rect.width() - 8,
                rect.bottom() - checkbox_rect.bottom() - 8,
            )
            if text_rect.height() > 0 and header_text:
                painter.drawText(
                    text_rect,
                    Qt.AlignHCenter | Qt.AlignVCenter | Qt.TextWordWrap,
                    header_text,
                )
        else:
            top_padding = 4
            if decoration_pixmap is not None and not decoration_pixmap.isNull():
                top_padding = max(top_padding, decoration_pixmap.height() + 6)
            text_rect = rect.adjusted(4, top_padding, -4, -4)
            if header_text:
                painter.drawText(
                    text_rect,
                    Qt.AlignHCenter | Qt.AlignVCenter | Qt.TextWordWrap,
                    header_text,
                )

        painter.restore()

    def sizeHint(self):
        hint = super().sizeHint()
        if hint.height() < 56:
            hint.setHeight(56)
        return hint

    def sectionSizeFromContents(self, logicalIndex):  # noqa: N802
        size = super().sectionSizeFromContents(logicalIndex)
        if size.height() < 56:
            size.setHeight(56)
        return size

    def _calculate_checkbox_rect(self, rect: QRect) -> QRect:
        option_button = QStyleOptionButton()
        option_button.state = QStyle.State_Enabled
        indicator = self.style().subElementRect(QStyle.SE_CheckBoxIndicator, option_button, self)
        checkbox_rect = QRect(indicator)
        center_point = QPoint(
            rect.center().x(),
            rect.top() + indicator.height() // 2 + 6,
        )
        checkbox_rect.moveCenter(center_point)
        return checkbox_rect


@dataclass
class Camada:
    """Representa uma camada do laminado."""

    idx: int
    material: str
    orientacao: Optional[float]
    ativo: bool
    simetria: bool
    ply_type: str = DEFAULT_PLY_TYPE
    ply_label: str = ""
    sequence: str = ""
    rosette: str = DEFAULT_ROSETTE_LABEL


@dataclass
class Laminado:
    """Agregado de metadados e camadas de um laminado."""

    nome: str
    tipo: str
    color_index: int | str = DEFAULT_COLOR_INDEX
    tag: str = ""
    celulas: list[str] = field(default_factory=list)
    camadas: list[Camada] = field(default_factory=list)
    auto_rename_enabled: bool = True


@dataclass
class GridModel:
    """Modelo raiz carregado da planilha do Grid Design."""

    laminados: Dict[str, Laminado] = field(default_factory=OrderedDict)
    celulas_ordenadas: list[str] = field(default_factory=list)
    cell_to_laminate: Dict[str, str] = field(default_factory=dict)
    compat_warnings: list[str] = field(default_factory=list)
    source_excel_path: Optional[str] = None
    dirty: bool = False

    def mark_dirty(self, value: bool = True) -> None:
        self.dirty = value

    def laminados_da_celula(self, cell_id: str) -> list[Laminado]:
        """Retorna laminados associados a uma celula."""
        mapped_name = self.cell_to_laminate.get(cell_id)
        if mapped_name:
            laminado = self.laminados.get(mapped_name)
            return [laminado] if laminado is not None else []
        return [
            laminado
            for laminado in self.laminados.values()
            if cell_id in laminado.celulas
        ]


def normalize_angle(value: object) -> float:
    """Normaliza a orientacao para graus decimais dentro do intervalo permitido."""
    if value is None:
        raise ValueError("orientacao ausente")

    if isinstance(value, bool):
        raise ValueError(f"valor booleano invalido para orientacao: {value!r}")

    number: float
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            raise ValueError("orientacao ausente")
        number = float(value)
    else:
        text = str(value).strip()
        if not text:
            raise ValueError("orientacao ausente")
        cleaned = text
        for token in _DEGREE_TOKENS:
            cleaned = cleaned.replace(token, "")
        cleaned = cleaned.replace("deg", "").replace("DEG", "").strip()
        cleaned = cleaned.replace(",", ".")
        if not _ORIENTATION_TEXT_PATTERN.fullmatch(cleaned):
            raise ValueError(f"orientacao invalida: {value!r}")
        try:
            number = float(cleaned)
        except ValueError as exc:  # pragma: no cover - defensive
            raise ValueError(f"orientacao invalida: {value!r}") from exc

    if number < ORIENTATION_MIN or number > ORIENTATION_MAX:
        raise ValueError(
            f"orientacao {number} fora do intervalo permitido [{ORIENTATION_MIN}, {ORIENTATION_MAX}]"
        )
    if math.isclose(number, 0.0, abs_tol=1e-9):
        return 0.0
    return number


def _normalized_orientation_token(value: object) -> Optional[float]:
    """Best-effort normalization without raising, returning None on failure."""
    try:
        return normalize_angle(value)
    except Exception:
        try:
            return normalize_angle(str(value))
        except Exception:
            return None


def layer_has_orientation(camada: Camada) -> bool:
    """Retorna True se a camada possui orientacao preenchida."""
    return getattr(camada, "orientacao", None) is not None


def count_oriented_layers(layers: Iterable[Camada]) -> int:
    """Conta apenas camadas com orientacao preenchida."""
    return sum(1 for camada in layers if layer_has_orientation(camada))


def orientation_highlight_color(value: object) -> Optional[QColor]:
    """Retorna a cor de destaque para uma orientacao ou None se nao aplicavel."""
    token = _normalized_orientation_token(value)
    if token is None:
        return None
    for key, color in ORIENTATION_HIGHLIGHT_COLORS.items():
        if math.isclose(token, key, abs_tol=1e-9):
            return color
    return DEFAULT_ORIENTATION_HIGHLIGHT


def format_orientation_value(value: Optional[float]) -> str:
    """Retorna a orientacao formatada com simbolo de grau."""
    if value is None:
        return ""
    number = float(value)
    if math.isclose(number, 0.0, abs_tol=1e-9):
        number = 0.0
    if number.is_integer():
        text = str(int(number))
    else:
        text = f"{number}".rstrip("0").rstrip(".")
        if not text or text in {"-0", "-0.0"}:
            text = "0"
    return f"{text}\N{DEGREE SIGN}"


def normalize_bool(value: object) -> bool:
    """Normaliza textos 'Sim/NAo' ou 'Yes/No' (e similares) em booleano."""
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return False
        return float(value) != 0.0

    text = str(value).strip()
    if not text:
        return False

    normalized = _normalize_header(text)
    true_values = {"yes", "y", "true", "1", "sim", "s"}
    false_values = {"no", "n", "false", "0", "nao"}

    if normalized in true_values:
        return True
    if normalized in false_values:
        return False

    logger.warning("Valor booleano desconhecido '%s'; assumindo False.", value)
    return False


def normalize_color_index(value: object, default: int = DEFAULT_COLOR_INDEX) -> int:
    """Normaliza o indice de cor (1-150) aceitando numeros inteiros."""
    if _is_blank(value):
        return default

    number: Optional[int] = None
    if isinstance(value, bool):
        logger.warning("Valor booleano invalido '%s' para indice de cor; usando %d.", value, default)
        return default

    if isinstance(value, (int, float)):
        if isinstance(value, float):
            if math.isnan(value):
                return default
            if not value.is_integer():
                logger.warning(
                    "Indice de cor nao inteiro '%s'; usando %d.",
                    value,
                    default,
                )
                return default
            number = int(value)
        else:
            number = int(value)
    else:
        text = str(value).strip()
        if not text:
            return default
        if re.fullmatch(r"#?[0-9a-fA-F]{6}", text):
            logger.warning(
                "Indice de cor em formato hexadecimal '%s' detectado; usando %d.",
                value,
                default,
            )
            return default
        try:
            as_float = float(text)
        except ValueError:
            logger.warning("Indice de cor invalido '%s'; usando %d.", value, default)
            return default
        if not as_float.is_integer():
            logger.warning(
                "Indice de cor nao inteiro '%s'; usando %d.",
                value,
                default,
            )
            return default
        number = int(as_float)

    if number is None:
        return default
    if MIN_COLOR_INDEX <= number <= MAX_COLOR_INDEX:
        return number
    logger.warning(
        "Indice de cor fora do intervalo (%d-%d): %s; usando %d.",
        MIN_COLOR_INDEX,
        MAX_COLOR_INDEX,
        number,
        default,
    )
    return default


def normalize_hex_color(value: object) -> Optional[str]:
    """Retorna cor hexadecimal no formato #RRGGBB se aplic├ível."""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    text = str(value).strip()
    if not text:
        return None
    candidate = text.replace("#", "").strip()
    if not _HEX_COLOR_PATTERN.fullmatch(candidate):
        return None
    return f"#{candidate.upper()}"


def load_grid_spreadsheet(path: str) -> GridModel:
    """Carrega a planilha do Grid Design em um GridModel."""
    file_path = Path(path)
    if not file_path.exists():
        raise ValueError(f"Arquivo '{path}' nAo encontrado.")

    ext = file_path.suffix.lower()
    if ext not in {".xls", ".xlsx"}:
        raise ValueError("Formato de planilha nAo suportado (use .xls ou .xlsx).")

    workbook = _open_workbook(file_path, ext)
    if "Planilha1" not in workbook.sheet_names:
        logger.error("Aba Planilha1 ausente.")
        raise ValueError(
            "A planilha deve conter a aba 'Planilha1' no formato exportado pelo Grid Design."
        )

    logger.info("Aba Planilha1 localizada - iniciando leitura.")
    df = _parse_sheet(workbook, "Planilha1")
    df = df.dropna(how="all").reset_index(drop=True)
    if df.empty:
        raise ValueError("Planilha1 nAo contAm dados para importar.")

    # Extrai celulas e mapeamento utilizando a nova funcao pAoblica.
    celulas_ordenadas = parse_cells_from_planilha1(df)
    cells_info = _extract_cells_section(df)
    cell_to_laminate = cells_info.mapping
    separator_idx = cells_info.separator_idx
    config_section = df.iloc[separator_idx + 1 :]

    if config_section.empty:
        raise ValueError("Planilha1: secao de configuracao de laminados estA vazia.")

    laminados = _parse_configuration_section(config_section)

    for cell_id, laminate_name in cell_to_laminate.items():
        laminado = laminados.get(laminate_name)
        if laminado is None:
            logger.warning(
                "CAlula '%s' associa laminado '%s', que nAo estA definido na secao de configuracao.",
                cell_id,
                laminate_name,
            )
            continue
        if cell_id not in laminado.celulas:
            laminado.celulas.append(cell_id)

    total_camadas = sum(
        count_oriented_layers(laminado.camadas) for laminado in laminados.values()
    )
    logger.info("Celulas importadas: %s", ", ".join(celulas_ordenadas))
    logger.info(
        "Planilha carregada com %d laminados, %d camadas e %d celulas.",
        len(laminados),
        total_camadas,
        len(celulas_ordenadas),
    )

    model = GridModel(
        laminados=laminados,
        celulas_ordenadas=celulas_ordenadas,
        cell_to_laminate=dict(cell_to_laminate),
    )
    model.source_excel_path = str(file_path)
    return model


def save_grid_spreadsheet(path: str, model: GridModel) -> None:
    """Persistir o GridModel no layout Planilha1 esperado pelo import."""
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows: list[list[object]] = []

    # Secao de celulas.
    rows.append(["Cells", "Laminate"])
    for cell_id in model.celulas_ordenadas:
        rows.append([cell_id, model.cell_to_laminate.get(cell_id, "")])
    rows.append(["#"])

    # Secao de configuracao por laminado.
    laminados_iter = model.laminados.values()
    for laminado in laminados_iter:
        rows.append(["Name", laminado.nome])
        color_value = laminado.color_index if laminado.color_index else DEFAULT_COLOR_INDEX
        rows.append(["ColorIdx", color_value])
        rows.append(["Type", laminado.tipo or ""])
        rows.append(["Stacking"])

        for camada in laminado.camadas:
            material_text = (camada.material or "").strip()
            has_orientation = camada.orientacao is not None
            if not material_text and not has_orientation:
                continue
            orientation_value = (
                camada.orientacao if camada.orientacao is not None else ""
            )
            rows.append(
                [
                    camada.material,
                    orientation_value,
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            )
        rows.append(["#"])

    if output_path.suffix.lower() == ".xls":
        try:
            import xlwt  # type: ignore
        except ImportError as exc:  # pragma: no cover - dependAancia opcional
            raise ValueError(
                "Salvar em '.xls' requer a dependAancia 'xlwt'. "
                "Instale 'xlwt' ou salve o arquivo com extensAo '.xlsx'."
            ) from exc

        book = xlwt.Workbook()
        sheet = book.add_sheet("Planilha1")
        for r_idx, row in enumerate(rows):
            for c_idx, value in enumerate(row):
                sheet.write(r_idx, c_idx, value)
        book.save(str(output_path))
        return

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Planilha1"
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(output_path)


class StackingTableModel(QAbstractTableModel):
    """Apresenta as camadas do laminado na tabela de stacking."""

    COL_NUMBER = 0
    COL_SELECT = 1
    COL_SEQUENCE = 2
    COL_PLY = 3
    COL_PLY_TYPE = 4
    COL_MATERIAL = 5
    COL_ORIENTATION = 6

    COL_NON_STRUCT = COL_PLY_TYPE  # Alias para compatibilidade retroativa.
    COL_CHECK = COL_SELECT

    HEADERS = [
        "#",
        "Selection",
        "Sequence",
        "Ply",
        "Simetria",
        "Material",
        "Orientacao",
    ]
    _LABEL_PATTERN = re.compile(r"^(?P<prefix>[A-Za-z][A-Za-z0-9_-]*)\.?(?P<number>\d+)$")
    headers = HEADERS

    def __init__(
        self,
        camadas: list[Camada] | None = None,
        change_callback: Optional[Callable[[list[Camada]], None]] = None,
        *,
        undo_stack: Optional[QUndoStack] = None,
        most_used_material_provider: Optional[Callable[[], Optional[str]]] = None,
    ) -> None:
        super().__init__()
        self._camadas: list[Camada] = []
        self._checked: list[bool] = []
        self._change_callback = change_callback
        self._rows_red: set[int] = set()
        self._rows_green: set[int] = set()
        self._undo_stack = undo_stack
        self._undo_suppressed = 0
        self._most_used_material_provider = most_used_material_provider or (lambda: None)
        self._unbalanced_warning = False
        self.update_layers(camadas or [])

    def set_undo_stack(self, undo_stack: Optional[QUndoStack]) -> None:
        self._undo_stack = undo_stack

    @contextmanager
    def suppress_undo(self) -> Iterable[None]:
        self._undo_suppressed += 1
        try:
            yield
        finally:
            self._undo_suppressed = max(0, self._undo_suppressed - 1)

    def _should_record_undo(self) -> bool:
        return self._undo_stack is not None and self._undo_suppressed == 0

    def _prefix_and_separator(self, attr: str, default_prefix: str) -> tuple[str, str]:
        for camada in self._camadas:
            text = str(getattr(camada, attr, "") or "").strip()
            match = self._LABEL_PATTERN.fullmatch(text)
            if match:
                prefix = match.group("prefix") or default_prefix
                separator = "." if "." in text else ""
                return prefix, separator
        return default_prefix, "."

    def _default_sequence_label(self, row: int) -> str:
        prefix, separator = self._prefix_and_separator("sequence", "Seq")
        return f"{prefix}{separator}{row + 1}"

    def _default_ply_label(self, row: int) -> str:
        prefix, separator = self._prefix_and_separator("ply_label", "Ply")
        return f"{prefix}{separator}{row + 1}"

    def _normalize_sequence_input(self, value: object, row: int) -> Optional[str]:
        text = str(value or "").strip()
        if not text:
            return self._default_sequence_label(row)
        match = self._LABEL_PATTERN.fullmatch(text)
        if match is None:
            return None
        prefix = match.group("prefix") or "Seq"
        try:
            number = int(match.group("number"))
        except ValueError:
            return None
        if number <= 0:
            return None
        separator = "." if "." in text else ""
        return f"{prefix}{separator}{number}"

    def _normalize_ply_input(self, value: object, row: int) -> Optional[str]:
        text = str(value or "").strip()
        if not text:
            return self._default_ply_label(row)
        match = self._LABEL_PATTERN.fullmatch(text)
        if match is None:
            return None
        prefix = match.group("prefix") or "Ply"
        try:
            number = int(match.group("number"))
        except ValueError:
            return None
        if number <= 0:
            return None
        separator = "." if "." in text else ""
        return f"{prefix}{separator}{number}"

    def _force_sequence_sync(self) -> None:
        if not self._camadas:
            return
        prefix, separator = self._prefix_and_separator("sequence", "Seq")
        changed_rows: list[int] = []
        for idx, camada in enumerate(self._camadas):
            expected = f"{prefix}{separator}{idx + 1}"
            if camada.sequence != expected:
                camada.sequence = expected
                changed_rows.append(idx)
        if not changed_rows:
            return
        for row in changed_rows:
            index = self.index(row, self.COL_SEQUENCE)
            self.dataChanged.emit(index, index, [Qt.DisplayRole, Qt.EditRole])

    def _force_ply_sync(self) -> None:
        if not self._camadas:
            return
        prefix, separator = self._prefix_and_separator("ply_label", "Ply")
        changed_rows: list[int] = []
        for idx, camada in enumerate(self._camadas):
            expected = f"{prefix}{separator}{idx + 1}"
            if camada.ply_label != expected:
                camada.ply_label = expected
                changed_rows.append(idx)
        if not changed_rows:
            return
        for row in changed_rows:
            index = self.index(row, self.COL_PLY)
            self.dataChanged.emit(index, index, [Qt.DisplayRole, Qt.EditRole])

    def _apply_field_value(self, row: int, column: int, value: object) -> bool:
        if not (0 <= row < len(self._camadas)):
            return False
        camada = self._camadas[row]
        extra_columns: list[int] = []
        if column == self.COL_SEQUENCE:
            normalized = self._normalize_sequence_input(value, row)
            if normalized is None:
                return False
            camada.sequence = normalized
        elif column == self.COL_PLY:
            normalized_ply = self._normalize_ply_input(value, row)
            if normalized_ply is None:
                return False
            camada.ply_label = normalized_ply
        elif column == self.COL_PLY_TYPE:
            normalized_type = normalize_ply_type_label(value)
            camada.ply_type = normalized_type
            material_text = str(getattr(camada, "material", "") or "")
            if (
                material_text
                and "foil" in material_text.lower()
                and hasattr(camada, "_auto_symmetry_backup")
            ):
                setattr(camada, "_auto_symmetry_backup", normalized_type)
        elif column == self.COL_MATERIAL:
            new_material = str(value or "").strip()
            if new_material and camada.orientacao is None:
                return False
            camada.material = new_material
            material_lower = new_material.lower()
            if material_lower and "foil" in material_lower:
                setattr(camada, "_auto_symmetry_backup", camada.ply_type)
                target_label = PLY_TYPE_OPTIONS[1]
                if camada.ply_type != target_label:
                    camada.ply_type = target_label
                    extra_columns.append(self.COL_PLY_TYPE)
            else:
                backup = getattr(camada, "_auto_symmetry_backup", None)
                if backup is not None:
                    if camada.ply_type != backup:
                        camada.ply_type = backup
                        extra_columns.append(self.COL_PLY_TYPE)
                    try:
                        delattr(camada, "_auto_symmetry_backup")
                    except AttributeError:
                        pass
        elif column == self.COL_ORIENTATION:
            old_orientation = getattr(camada, "orientacao", None)
            if value is None:
                camada.orientacao = None
            else:
                text = str(value).strip()
                if not text or text.lower() == "x":
                    camada.orientacao = None
                else:
                    try:
                        camada.orientacao = normalize_angle(value)
                    except (TypeError, ValueError):
                        try:
                            camada.orientacao = normalize_angle(text)
                        except (TypeError, ValueError):
                            return False
            if camada.orientacao is None and camada.material:
                camada.material = ""
                extra_columns.append(self.COL_MATERIAL)
                if camada.ply_type != PLY_TYPE_OPTIONS[1]:
                    camada.ply_type = PLY_TYPE_OPTIONS[1]
                    extra_columns.append(self.COL_PLY_TYPE)
            elif (
                camada.orientacao is not None
                and not camada.material
                and callable(self._most_used_material_provider)
            ):
                suggestion = str(self._most_used_material_provider() or "").strip()
                if suggestion:
                    camada.material = suggestion
                    extra_columns.append(self.COL_MATERIAL)
                if old_orientation is None and camada.ply_type == PLY_TYPE_OPTIONS[1]:
                    camada.ply_type = PLY_TYPE_OPTIONS[0]
                    extra_columns.append(self.COL_PLY_TYPE)
        else:
            return False

        index = self.index(row, column)
        self.dataChanged.emit(index, index, [Qt.DisplayRole, Qt.EditRole])
        for extra_column in extra_columns:
            extra_index = self.index(row, extra_column)
            self.dataChanged.emit(extra_index, extra_index, [Qt.DisplayRole, Qt.EditRole])
        if self._change_callback:
            self._change_callback(self.layers())
        return True

    def apply_field_value(self, row: int, column: int, value: object) -> bool:
        with self.suppress_undo():
            return self._apply_field_value(row, column, value)

    def _apply_or_record_change(
        self, row: int, column: int, old_value: object, new_value: object
    ) -> bool:
        if self._should_record_undo() and self._undo_stack is not None:
            description = self._describe_edit(column, row)
            command = _LayerFieldEditCommand(
                self,
                row,
                column,
                old_value,
                new_value,
                description,
            )
            self._undo_stack.push(command)
            return True
        return self._apply_field_value(row, column, new_value)

    def _describe_edit(self, column: int, row: int) -> str:
        labels = {
            self.COL_SEQUENCE: "sequencia",
            self.COL_PLY: "ply",
            self.COL_PLY_TYPE: "simetria",
            self.COL_MATERIAL: "material",
            self.COL_ORIENTATION: "orienta\u00e7\u00e3o",
        }
        label = labels.get(column, "camada")
        return f"Atualizar {label} (linha {row + 1})"

    def update_layers(self, camadas: Iterable[Camada]) -> None:
        self.beginResetModel()
        self._camadas = [self._ensure_layer_defaults(c) for c in camadas]
        self._checked = [False] * len(self._camadas)
        self._rows_red.clear()
        self._rows_green.clear()
        self._sync_indices()
        self._force_sequence_sync()
        self._force_ply_sync()
        self.endResetModel()

    def set_unbalanced_warning(self, enabled: bool) -> None:
        enabled = bool(enabled)
        if self._unbalanced_warning == enabled:
            return
        self._unbalanced_warning = enabled
        self.headerDataChanged.emit(Qt.Horizontal, self.COL_ORIENTATION, self.COL_ORIENTATION)

    @staticmethod
    def _ensure_layer_defaults(camada: Camada) -> Camada:
        legacy_flag = getattr(camada, "nao_estrutural", False)
        normalized_label = normalize_ply_type_label(getattr(camada, "ply_type", None))
        if legacy_flag and normalized_label == DEFAULT_PLY_TYPE:
            normalized_label = PLY_TYPE_OPTIONS[1]
        camada.ply_type = normalized_label
        camada.ply_label = str(getattr(camada, "ply_label", "") or "")
        rosette_value = str(getattr(camada, "rosette", "") or "").strip()
        camada.rosette = rosette_value or DEFAULT_ROSETTE_LABEL
        if getattr(camada, "orientacao", None) is None:
            camada.material = ""
            camada.ply_type = PLY_TYPE_OPTIONS[1]
        if hasattr(camada, "nao_estrutural"):
            try:
                delattr(camada, "nao_estrutural")
            except AttributeError:
                pass
        return camada

    def rowCount(self, parent: QModelIndex = QModelIndex()) -> int:  # noqa: N802
        if parent.isValid():
            return 0
        return len(self._camadas)

    def columnCount(self, parent: QModelIndex = QModelIndex()) -> int:  # noqa: N802
        if parent.isValid():
            return 0
        return len(self.HEADERS)

    def headerData(  # noqa: N802
        self,
        section: int,
        orientation: Qt.Orientation,
        role: int = Qt.DisplayRole,
    ) -> Optional[str]:
        if orientation == Qt.Horizontal:
            if role == Qt.DecorationRole and section == self.COL_ORIENTATION:
                if self._unbalanced_warning:
                    return QApplication.style().standardIcon(QStyle.SP_MessageBoxWarning)
                return None
            if role == Qt.DisplayRole and 0 <= section < len(self.HEADERS):
                return self.HEADERS[section]
            if role == Qt.TextAlignmentRole:
                return int(Qt.AlignCenter)
            if role == Qt.ToolTipRole and 0 <= section < len(self.HEADERS):
                return self.HEADERS[section]
        if orientation == Qt.Vertical and role == Qt.DisplayRole:
            return str(section + 1)
        return super().headerData(section, orientation, role)

    def data(self, index: QModelIndex, role: int = Qt.DisplayRole) -> Optional[object]:  # noqa: N802
        if not index.isValid() or not (0 <= index.row() < len(self._camadas)):
            return None

        camada = self._camadas[index.row()]
        column = index.column()

        if role == Qt.DisplayRole:
            if column == self.COL_NUMBER:
                return str(index.row() + 1)
            if column == self.COL_SEQUENCE:
                return camada.sequence or self._default_sequence_label(index.row())
            if column == self.COL_PLY:
                return camada.ply_label or self._default_ply_label(index.row())
            if column == self.COL_PLY_TYPE:
                return normalize_ply_type_label(camada.ply_type)
            if column == self.COL_MATERIAL:
                return camada.material
            if column == self.COL_ORIENTATION:
                if camada.orientacao is None:
                    return "Empty"
                return format_orientation_value(camada.orientacao)
        elif role == Qt.EditRole:
            if column == self.COL_SEQUENCE:
                return camada.sequence or self._default_sequence_label(index.row())
            if column == self.COL_PLY:
                return camada.ply_label or self._default_ply_label(index.row())
            if column == self.COL_PLY_TYPE:
                return normalize_ply_type_label(camada.ply_type)
            if column == self.COL_MATERIAL:
                return camada.material
            if column == self.COL_ORIENTATION:
                if camada.orientacao is None:
                    return ""
                try:
                    return f"{float(camada.orientacao):g}"
                except Exception:
                    return str(camada.orientacao)
        elif role == Qt.CheckStateRole:
            if column == self.COL_SELECT:
                return Qt.Checked if self._checked[index.row()] else Qt.Unchecked
        elif role == Qt.TextAlignmentRole:
            if column in (
                self.COL_NUMBER,
                self.COL_SELECT,
                self.COL_SEQUENCE,
                self.COL_PLY,
                self.COL_PLY_TYPE,
                self.COL_ORIENTATION,
            ):
                return int(Qt.AlignVCenter | Qt.AlignCenter)
            return int(Qt.AlignVCenter | Qt.AlignLeft)
        elif role == Qt.ForegroundRole:
            if column == self.COL_ORIENTATION and camada.orientacao is None:
                return QColor(160, 160, 160)
        elif role == Qt.BackgroundRole:
            row = index.row()
            if row in self._rows_red:
                return QColor(220, 53, 69)
            if row in self._rows_green:
                return QColor(40, 167, 69)
            if column == self.COL_ORIENTATION:
                color = orientation_highlight_color(camada.orientacao)
                if color is not None:
                    return color
        elif role == ORIENTATION_SYMMETRY_ROLE:
            if column == self.COL_ORIENTATION and index.row() in self._rows_green:
                return True

        return None

    def _emit_rows_changed(self, rows: Iterable[int]) -> None:
        row_list = [r for r in sorted(set(rows)) if 0 <= r < self.rowCount()]
        if not row_list:
            return
        if self.columnCount() == 0:
            return
        top = row_list[0]
        bottom = row_list[-1]
        left = self.index(top, 0)
        right = self.index(bottom, self.columnCount() - 1)
        self.dataChanged.emit(left, right, [Qt.BackgroundRole])

    def clear_all_highlights(self) -> None:
        if not (self._rows_red or self._rows_green):
            return
        rows = list(self._rows_red | self._rows_green)
        self._rows_red.clear()
        self._rows_green.clear()
        self._emit_rows_changed(rows)

    def set_red_rows(self, rows: Iterable[int]) -> None:
        valid_rows = {r for r in rows if 0 <= r < self.rowCount()}
        if self._rows_red == valid_rows:
            return
        self._rows_red = valid_rows
        self._emit_rows_changed(valid_rows)

    def add_red_rows(self, rows: Iterable[int]) -> None:
        new_rows = {r for r in rows if 0 <= r < self.rowCount()}
        if not new_rows:
            return
        before = set(self._rows_red)
        self._rows_red |= new_rows
        self._emit_rows_changed(self._rows_red | before)

    def set_green_rows(self, rows: Iterable[int]) -> None:
        valid_rows = {r for r in rows if 0 <= r < self.rowCount()}
        if self._rows_green == valid_rows:
            return
        self._rows_green = valid_rows
        self._emit_rows_changed(valid_rows)

    def add_green_rows(self, rows: Iterable[int]) -> None:
        new_rows = {r for r in rows if 0 <= r < self.rowCount()}
        if not new_rows:
            return
        before = set(self._rows_green)
        self._rows_green |= new_rows
        self._emit_rows_changed(self._rows_green | before)

    def _shift_highlights_on_insert(self, position: int) -> None:
        if not self._rows_red and not self._rows_green:
            return
        changed: set[int] = set()

        def shift(rows: set[int]) -> set[int]:
            updated: set[int] = set()
            for row in rows:
                if row >= position:
                    changed.add(row)
                    new_row = row + 1
                    changed.add(new_row)
                    updated.add(new_row)
                else:
                    updated.add(row)
            return updated

        self._rows_red = shift(self._rows_red)
        self._rows_green = shift(self._rows_green)
        if changed:
            self._emit_rows_changed(changed)

    def _adjust_highlights_on_remove(self, removed_rows: Iterable[int]) -> None:
        removed_sorted = sorted({r for r in removed_rows if r >= 0})
        if not removed_sorted:
            return
        removed_set = set(removed_sorted)
        changed: set[int] = set()

        def adjust(rows: set[int]) -> set[int]:
            updated: set[int] = set()
            for row in rows:
                if row in removed_set:
                    changed.add(row)
                    continue
                shift = sum(1 for r in removed_sorted if r < row)
                new_row = row - shift
                if new_row != row:
                    changed.add(row)
                    changed.add(new_row)
                updated.add(new_row)
            return updated

        self._rows_red = adjust(self._rows_red)
        self._rows_green = adjust(self._rows_green)
        if changed:
            self._emit_rows_changed(changed)

    def _update_highlights_on_move(self, source: int, target: int) -> None:
        if not self._rows_red and not self._rows_green:
            return
        changed: set[int] = set()

        def adjust(rows: set[int]) -> set[int]:
            updated: set[int] = set()
            for row in rows:
                if row == source:
                    changed.add(row)
                    changed.add(target)
                    updated.add(target)
                elif source < target and source < row <= target:
                    changed.add(row)
                    new_row = row - 1
                    changed.add(new_row)
                    updated.add(new_row)
                elif source > target and target <= row < source:
                    changed.add(row)
                    new_row = row + 1
                    changed.add(new_row)
                    updated.add(new_row)
                else:
                    updated.add(row)
            return updated

        self._rows_red = adjust(self._rows_red)
        self._rows_green = adjust(self._rows_green)
        if changed:
            self._emit_rows_changed(changed)

    def flags(self, index: QModelIndex) -> Qt.ItemFlags:  # noqa: N802
        if not index.isValid():
            return Qt.NoItemFlags
        column = index.column()
        base_flags = Qt.ItemIsSelectable | Qt.ItemIsEnabled
        if column == self.COL_SELECT:
            return base_flags | Qt.ItemIsUserCheckable
        if column in (self.COL_SEQUENCE, self.COL_PLY):
            return base_flags | Qt.ItemIsEditable
        if column in (self.COL_PLY_TYPE, self.COL_MATERIAL, self.COL_ORIENTATION):
            return base_flags | Qt.ItemIsEditable
        return base_flags

    def setData(  # noqa: N802
        self,
        index: QModelIndex,
        value: object,
        role: int = Qt.EditRole,
    ) -> bool:
        if not index.isValid():
            return False
        column = index.column()
        row = index.row()
        camada = self._camadas[row]

        if column == self.COL_SELECT and role == Qt.CheckStateRole:
            checked = value == Qt.Checked
            if self._checked[row] == checked:
                return False
            self._checked[row] = checked
            self.dataChanged.emit(index, index, [Qt.CheckStateRole])
            self.headerDataChanged.emit(Qt.Horizontal, self.COL_SELECT, self.COL_SELECT)
            return True

        if column == self.COL_PLY_TYPE and role in (Qt.EditRole, Qt.DisplayRole):
            text = normalize_ply_type_label(value)
            if camada.ply_type == text:
                return False
            return self._apply_or_record_change(row, column, camada.ply_type, text)

        if column == self.COL_SEQUENCE and role in (Qt.EditRole, Qt.DisplayRole):
            normalized = self._normalize_sequence_input(value, row)
            if normalized is None:
                return False
            if camada.sequence == normalized:
                return False
            return self._apply_or_record_change(row, column, camada.sequence, normalized)

        if column == self.COL_PLY and role in (Qt.EditRole, Qt.DisplayRole):
            normalized_ply = self._normalize_ply_input(value, row)
            if normalized_ply is None:
                return False
            if camada.ply_label == normalized_ply:
                return False
            return self._apply_or_record_change(
                row, column, camada.ply_label, normalized_ply
            )

        if column == self.COL_MATERIAL and role in (Qt.EditRole, Qt.DisplayRole):
            new_value = str(value).strip()
            if new_value == camada.material:
                return False
            return self._apply_or_record_change(row, column, camada.material, new_value)

        if column == self.COL_ORIENTATION and role in (Qt.EditRole, Qt.DisplayRole):
            previous_orientation = getattr(camada, "orientacao", None)
            if value is None:
                if previous_orientation is None:
                    return False
                return self._apply_or_record_change(row, column, previous_orientation, None)
            text = str(value).strip()
            if not text or text.lower() in {"x", "empty"}:
                if previous_orientation is None:
                    return False
                return self._apply_or_record_change(row, column, previous_orientation, None)
            try:
                angle = normalize_angle(value)
            except ValueError:
                try:
                    angle = normalize_angle(text)
                except ValueError:
                    return False
            if previous_orientation is not None and math.isclose(previous_orientation, angle, rel_tol=0.0, abs_tol=1e-9):
                return False
            return self._apply_or_record_change(row, column, previous_orientation, angle)

        return False
    # Helpers for controller ------------------------------------------------- #

    def insert_layer(self, position: int, camada: Camada) -> None:
        position = max(0, min(position, len(self._camadas)))
        self.beginInsertRows(QModelIndex(), position, position)
        self._camadas.insert(position, self._ensure_layer_defaults(camada))
        self._checked.insert(position, False)
        inserted = self._camadas[position]
        material_filled = False
        if inserted.orientacao is not None and not inserted.material:
            suggestion = str(self._most_used_material_provider() or "").strip()
            if suggestion:
                inserted.material = suggestion
                material_filled = True
        self.endInsertRows()
        self._shift_highlights_on_insert(position)
        self._sync_indices()
        self._force_sequence_sync()
        self._force_ply_sync()
        if material_filled:
            material_index = self.index(position, self.COL_MATERIAL)
            self.dataChanged.emit(material_index, material_index, [Qt.DisplayRole, Qt.EditRole])
        if self._change_callback:
            self._change_callback(self.layers())

    def remove_rows(self, rows: Iterable[int]) -> int:
        removed = 0
        removed_rows: list[int] = []
        for row in sorted(set(rows), reverse=True):
            if 0 <= row < len(self._camadas):
                self.beginRemoveRows(QModelIndex(), row, row)
                del self._camadas[row]
                del self._checked[row]
                self.endRemoveRows()
                removed += 1
                removed_rows.append(row)
        if removed:
            removed_rows.sort()
            self._adjust_highlights_on_remove(removed_rows)
            self._sync_indices()
            self._force_sequence_sync()
            self._force_ply_sync()
            if self._change_callback:
                self._change_callback(self.layers())
        return removed

    def checked_rows(self) -> list[int]:
        return [idx for idx, checked in enumerate(self._checked) if checked]

    def clear_checks(self) -> None:
        if not any(self._checked):
            return
        self.set_all_checked(False)

    def layers(self) -> list[Camada]:
        return list(self._camadas)

    def _sync_indices(self) -> None:
        for idx, camada in enumerate(self._camadas):
            camada.idx = idx

    def set_all_checked(self, value: bool) -> None:
        if not self._checked:
            return
        new_state = [value] * len(self._checked)
        if self._checked == new_state:
            return
        self._checked = new_state
        if self.rowCount() == 0:
            return
        top_left = self.index(0, self.COL_SELECT)
        bottom_right = self.index(self.rowCount() - 1, self.COL_SELECT)
        self.dataChanged.emit(top_left, bottom_right, [Qt.CheckStateRole])
        self.headerDataChanged.emit(Qt.Horizontal, self.COL_SELECT, self.COL_SELECT)

    def all_checked(self) -> bool:
        return bool(self._checked) and all(self._checked)

    def any_checked(self) -> bool:
        return any(self._checked)

    def move_row(self, source: int, target: int) -> bool:
        if source == target or not (0 <= source < len(self._camadas)):
            return False
        if not (0 <= target < len(self._camadas)):
            return False
        destination = target
        if destination > source:
            destination += 1
        self.beginMoveRows(QModelIndex(), source, source, QModelIndex(), destination)
        camada = self._camadas.pop(source)
        checked = self._checked.pop(source)
        self._camadas.insert(target, camada)
        self._checked.insert(target, checked)
        self.endMoveRows()
        self._update_highlights_on_move(source, target)
        self._sync_indices()
        self._force_sequence_sync()
        self._force_ply_sync()
        if self._change_callback:
            self._change_callback(self.layers())
        return True

    def toggle_check(self, row: int) -> bool:
        if not (0 <= row < len(self._camadas)):
            return False
        index = self.index(row, self.COL_SELECT)
        current = self.data(index, Qt.CheckStateRole)
        new_state = Qt.Unchecked if current == Qt.Checked else Qt.Checked
        return self.setData(index, new_state, Qt.CheckStateRole)


def bind_model_to_ui(model: GridModel, ui) -> None:
    """Efetua o binding do modelo carregado com os widgets da UI."""
    binding = getattr(ui, "_grid_binding", None)
    if binding is not None:
        binding.teardown()
    ui._grid_binding = _GridUiBinding(model, ui)  # type: ignore[attr-defined]


def bind_cells_to_ui(model: GridModel, ui) -> None:
    """
    Popula o listbox de celulas com base em ``model.celulas_ordenadas``.
    """
    list_widget = getattr(ui, "lstCelulas", None)
    if not isinstance(list_widget, QListWidget):
        list_widget = getattr(ui, "cells_list", None)
    if not isinstance(list_widget, QListWidget):
        return

    list_widget.blockSignals(True)
    list_widget.clear()
    for cell_id in model.celulas_ordenadas:
        item = QListWidgetItem(_format_cell_label(model, cell_id))
        item.setData(Qt.UserRole, cell_id)
        list_widget.addItem(item)
    list_widget.blockSignals(False)

    if model.celulas_ordenadas:
        list_widget.setCurrentRow(0)


def _format_cell_label(model: GridModel, cell_id: str) -> str:
    laminate_name = model.cell_to_laminate.get(cell_id) or ""
    if laminate_name:
        laminate = model.laminados.get(laminate_name)
        laminate_name = laminate.nome if laminate is not None else laminate_name
    else:
        laminados = model.laminados_da_celula(cell_id)
        laminate_name = laminados[0].nome if laminados else NO_LAMINATE_LABEL
    return f"{cell_id} | {laminate_name}"


# Helpers ------------------------------------------------------------------ #


class _WorkbookProtocol(Protocol):
    sheet_names: list[str]

    def parse(self, sheet_name: str, *args, **kwargs) -> pd.DataFrame:
        ...


def _open_workbook(file_path: Path, ext: str) -> _WorkbookProtocol:
    if ext == ".xls":
        if xlrd is None:
            raise ValueError(
                "Leitura de arquivos .xls requer a dependAancia 'xlrd==1.2.0'."
            )
        return _XlrdWorkbook(file_path)

    try:
        workbook = pd.ExcelFile(file_path, engine="openpyxl")
    except _OPEN_EXCEL_EXCEPTIONS as exc:
        return _open_with_xlrd(file_path, exc)
    except ValueError as exc:
        if "not a zip file" in str(exc).lower():
            return _open_with_xlrd(file_path, exc)
        raise ValueError(f"NAo foi possAvel abrir '{file_path}': {exc}") from exc
    except Exception as exc:  # pragma: no cover - protecao
        raise ValueError(f"NAo foi possAvel abrir '{file_path}': {exc}") from exc
    return workbook  # type: ignore[return-value]


def _parse_sheet(workbook: _WorkbookProtocol, sheet_name: str) -> pd.DataFrame:
    parse = getattr(workbook, "parse")
    try:
        return parse(sheet_name, header=None, dtype=object)
    except TypeError:
        return parse(sheet_name)


def _normalize_header(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_only = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    ascii_only = ascii_only.lower()
    return re.sub(r"[^a-z0-9]+", "", ascii_only)


def _find_separator_row(df: pd.DataFrame) -> Optional[int]:
    for idx, row in df.iterrows():
        for value in row:
            if _is_blank(value):
                continue
            if _is_separator_token(value):
                return idx
            break
    return None


def _is_separator_token(value: object) -> bool:
    text = str(value).strip().lower()
    text = text.replace(" ", "")
    return text == "#"


def _find_cells_header_row(df: pd.DataFrame, separator_idx: int) -> Optional[int]:
    normalized_aliases = {_normalize_header(alias) for alias in CELL_MAPPING_ALIASES}
    for idx in range(separator_idx):
        row = df.iloc[idx]
        found_cells = False
        for value in row:
            if _is_blank(value):
                continue
            normalized = _normalize_header(str(value))
            if normalized in normalized_aliases:
                found_cells = True
                break
        if found_cells:
            return idx
    return None


def _first_non_blank_value(row: pd.Series) -> Optional[str]:
    for value in row:
        if _is_blank(value):
            continue
        return str(value)
    return None


def _extract_cells_section(df: pd.DataFrame) -> _CellsSection:
    """Localiza a secao de celulas em Planilha1 e retorna dados estruturados."""
    separator_idx = _find_separator_row(df)
    if separator_idx is None:
        raise ValueError(
            "NAo foi possAvel localizar a linha separadora '#' em Planilha1."
        )

    cells_header_idx = _find_cells_header_row(df, separator_idx)
    if cells_header_idx is None:
        raise ValueError("NAo foi possAvel localizar a secao 'Cells' em Planilha1.")

    header_row = df.iloc[cells_header_idx]
    header_keys = [
        _normalize_header(str(value)) if not _is_blank(value) else ""
        for value in header_row
    ]

    cell_col = _find_column(header_keys, CELL_MAPPING_ALIASES)
    laminate_col = _find_column(header_keys, LAMINATE_ALIASES)
    data_start = cells_header_idx + 1

    if cell_col is None:
        cell_col = 0

    cells_ordered: list[str] = []
    seen_cells: set[str] = set()
    cell_to_laminate: Dict[str, str] = {}

    for row_idx in range(data_start, separator_idx):
        row = df.iloc[row_idx]
        cell_value = row.iloc[cell_col] if cell_col < len(row) else None
        if _is_blank(cell_value):
            continue
        cell_id = str(cell_value).strip().upper()
        if not CELL_ID_PATTERN.match(cell_id):
            logger.warning(
                "Linha %d da secao 'Cells': celula '%s' invalida, ignorada.",
                row_idx + 1,
                cell_value,
            )
            continue
        if cell_id not in seen_cells:
            cells_ordered.append(cell_id)
            seen_cells.add(cell_id)

        if laminate_col is not None and laminate_col < len(row):
            laminate_value = row.iloc[laminate_col]
            if not _is_blank(laminate_value) and cell_id not in cell_to_laminate:
                cell_to_laminate[cell_id] = str(laminate_value).strip()

    if not cells_ordered:
        raise ValueError("NAo hA celulas vAlidas entre 'Cells' e '#' em Planilha1.")

    logger.info(
        "Secao 'Cells' processada com %d celulas distintas.",
        len(cells_ordered),
    )
    return _CellsSection(cells=cells_ordered, mapping=cell_to_laminate, separator_idx=separator_idx)


def _parse_configuration_section(
    df: pd.DataFrame,
) -> OrderedDict[str, Laminado]:
    df = df.reset_index(drop=True)
    laminados: OrderedDict[str, Laminado] = OrderedDict()

    normalized_color_aliases = {_normalize_header(alias) for alias in COLOR_ALIASES}
    normalized_type_aliases = {_normalize_header(alias) for alias in TYPE_ALIASES}
    normalized_tag_aliases = {_normalize_header(alias) for alias in TAG_ALIASES}

    idx = 0
    while idx < len(df):
        row = df.iloc[idx]
        label_raw = row.iloc[0] if len(row) > 0 else None
        if _is_blank(label_raw):
            idx += 1
            continue

        if _is_separator_token(label_raw):
            idx += 1
            continue

        label = str(label_raw).strip()
        normalized_label = _normalize_header(label)
        if normalized_label != "name":
            idx += 1
            continue

        name_value = row.iloc[1] if len(row) > 1 else None
        if _is_blank(name_value):
            raise ValueError(f"Linha {idx + 1}: laminado sem nome definido.")
        laminate_name = str(name_value).strip()
        idx += 1

        color_index = DEFAULT_COLOR_INDEX
        laminate_type = ""
        laminate_tag = ""

        while idx < len(df):
            row = df.iloc[idx]
            first_val = row.iloc[0] if len(row) > 0 else None

            if _is_blank(first_val):
                idx += 1
                continue

            if _is_separator_token(first_val):
                break

            normalized = _normalize_header(str(first_val))
            if normalized == "name":
                break

            if normalized in normalized_color_aliases:
                color_value = row.iloc[1] if len(row) > 1 else None
                hex_color = normalize_hex_color(color_value)
                if hex_color:
                    color_index = hex_color
                else:
                    color_index = normalize_color_index(color_value)
                idx += 1
                continue

            if normalized in normalized_type_aliases:
                laminate_type = _string_or_empty(row.iloc[1] if len(row) > 1 else "")
                idx += 1
                continue

            if normalized in normalized_tag_aliases:
                laminate_tag = _string_or_empty(row.iloc[1] if len(row) > 1 else "")
                idx += 1
                continue

            if normalized == "stacking":
                idx += 1
                break

            logger.warning(
                "Linha %d: rA3tulo '%s' desconhecido na configuracao de laminados; ignorando.",
                idx + 1,
                first_val,
            )
            idx += 1

        layers: list[Camada] = []
        while idx < len(df):
            row = df.iloc[idx]
            first_val = row.iloc[0] if len(row) > 0 else None

            if not _is_blank(first_val) and (
                _is_separator_token(first_val)
                or _normalize_header(str(first_val)) == "name"
            ):
                break

            if _is_blank(first_val):
                idx += 1
                continue

            material = str(first_val).strip()
            orientation_raw = row.iloc[1] if len(row) > 1 else 0
            try:
                orientation = normalize_angle(orientation_raw)
            except ValueError as exc:
                raise ValueError(
                    f"Linha {idx + 1}: laminado '{laminate_name}' possui orientacao invalida ({exc})."
                ) from exc

            active_raw = row.iloc[2] if len(row) > 2 else None
            symmetry_raw = row.iloc[3] if len(row) > 3 else None
            active = normalize_bool(active_raw) if not _is_blank(active_raw) else True
            symmetry = (
                normalize_bool(symmetry_raw) if not _is_blank(symmetry_raw) else False
            )
            ply_raw: object
            if len(row) > 6 and not _is_blank(row.iloc[6]):
                ply_raw = row.iloc[6]
            else:
                ply_raw = row.iloc[4] if len(row) > 4 else None
            if _is_blank(ply_raw):
                ply_type_value = DEFAULT_PLY_TYPE
            else:
                candidate = str(ply_raw).strip()
                if not is_known_ply_type_value(candidate):
                    logger.warning(
                        "Laminado '%s': valor de Simetria '%s' invalido; usando '%s'.",
                        laminate_name,
                        candidate,
                        DEFAULT_PLY_TYPE,
                    )
                ply_type_value = normalize_ply_type_label(candidate)

            layers.append(
                Camada(
                    idx=len(layers),
                    material=material,
                    orientacao=orientation,
                    ativo=active,
                    simetria=symmetry,
                    ply_type=ply_type_value,
                    ply_label=f"Ply.{len(layers) + 1}",
                    sequence=f"Seq.{len(layers) + 1}",
                    rosette=DEFAULT_ROSETTE_LABEL,
                )
            )
            idx += 1

        if not layers:
            logger.warning(
                "Laminado '%s' nAo possui camadas definidas na secao de stacking.",
                laminate_name,
            )

        laminados[laminate_name] = Laminado(
            nome=laminate_name,
            tipo=laminate_type,
            color_index=color_index,
            tag=laminate_tag,
            celulas=[],
            camadas=layers,
        )

    if not laminados:
        raise ValueError(
            "Planilha1: secao de configuracao nAo definiu nenhum laminado."
        )

    return laminados


def _find_column(header_keys: list[str], aliases: Iterable[str]) -> Optional[int]:
    normalized_aliases = {_normalize_header(alias) for alias in aliases}
    for idx, key in enumerate(header_keys):
        if key in normalized_aliases:
            return idx
    return None


def _require_column(
    header_keys: list[str], aliases: Iterable[str], display_name: str
) -> int:
    column = _find_column(header_keys, aliases)
    if column is None:
        raise ValueError(
            f"Coluna obrigatA3ria '{display_name}' ausente em Planilha1 (secao de configuracao abaixo de '#')."
        )
    return column


class _GridUiBinding:
    """Encapsula o binding entre GridModel e widgets do MainWindow."""

    def __init__(self, model: GridModel, ui) -> None:
        self.model = model
        self.ui = ui
        self._updating = False
        self._current_laminate: Optional[str] = None
        self._current_cell_id: Optional[str] = None
        self._laminates_by_cell: dict[str, list[str]] = self._build_cell_index()
        undo_stack = getattr(self.ui, "undo_stack", None)
        self.stacking_model = StackingTableModel(
            change_callback=self._on_layers_modified,
            undo_stack=undo_stack,
            most_used_material_provider=lambda: self._most_used_material(),
        )
        self._cells_widget: Optional[QListWidget] = None
        self._table_view: Optional[QTableView] = None
        self._header_view: Optional[WordWrapHeader] = None

        self.stacking_model.dataChanged.connect(lambda *args: self._update_layers_count())
        self.stacking_model.rowsInserted.connect(self._update_layers_count)
        self.stacking_model.rowsRemoved.connect(self._update_layers_count)
        self.stacking_model.modelReset.connect(self._update_layers_count)
        self.stacking_model.dataChanged.connect(self._refresh_selection_header)
        self.stacking_model.rowsInserted.connect(self._refresh_selection_header)
        self.stacking_model.rowsRemoved.connect(self._refresh_selection_header)
        self.stacking_model.modelReset.connect(self._refresh_selection_header)

        self._setup_widgets()
        self._connect_signals()
        self._update_layers_count()

    def teardown(self) -> None:
        """Remove conexAes quando um novo binding for aplicado."""
        try:
            cells_widget = getattr(self.ui, "lstCelulas", None)
            if cells_widget is None:
                cells_widget = getattr(self.ui, "cells_list", None)
            if isinstance(cells_widget, QListWidget):
                cells_widget.currentItemChanged.disconnect(self._on_cell_item_changed)
        except (AttributeError, TypeError):
            pass
        try:
            self.ui.laminate_name_combo.currentTextChanged.disconnect(self._on_laminate_selected)
        except (AttributeError, TypeError):
            pass
        try:
            if isinstance(getattr(self, "_table_view", None), QTableView):
                self._table_view.clicked.disconnect(self._on_table_clicked)
        except (AttributeError, TypeError):
            pass
        for signal in (
            self.stacking_model.dataChanged,
            self.stacking_model.rowsInserted,
            self.stacking_model.rowsRemoved,
            self.stacking_model.modelReset,
        ):
            try:
                signal.disconnect(self._refresh_selection_header)
            except (AttributeError, TypeError):
                pass
        self._table_view = None
        self.set_header_view(None)

    # Internal helpers -------------------------------------------------- #

    def _build_cell_index(self) -> dict[str, list[str]]:
        if self.model.cell_to_laminate:
            mapping: dict[str, list[str]] = {}
            for cell, lam in self.model.cell_to_laminate.items():
                if lam:
                    mapping.setdefault(cell, []).append(lam)
            return mapping
        mapping: dict[str, list[str]] = {}
        for name, laminado in self.model.laminados.items():
            for cell_id in laminado.celulas:
                mapping.setdefault(cell_id, []).append(name)
        return mapping

    def _setup_widgets(self) -> None:
        name_combo = getattr(self.ui, "laminate_name_combo", None)
        if isinstance(name_combo, QComboBox):
            name_combo.blockSignals(True)
            name_combo.clear()
            for laminado in self.model.laminados.values():
                name_combo.addItem(laminado.nome)
            name_combo.blockSignals(False)

        color_combo = getattr(self.ui, "laminate_color_combo", None)
        if isinstance(color_combo, QComboBox):
            color_combo.blockSignals(True)
            color_combo.clear()
            color_combo.addItems([str(idx) for idx in range(MIN_COLOR_INDEX, MAX_COLOR_INDEX + 1)])
            color_combo.setEditable(False)
            color_combo.blockSignals(False)

        type_combo = getattr(self.ui, "laminate_type_combo", None)
        if isinstance(type_combo, QComboBox):
            type_combo.blockSignals(True)
            unique_types = [
                laminado.tipo for laminado in self.model.laminados.values() if laminado.tipo
            ]
            for tipo in OrderedDict((item, None) for item in unique_types):
                if tipo not in [type_combo.itemText(i) for i in range(type_combo.count())]:
                    type_combo.addItem(tipo)
            type_combo.blockSignals(False)

        table = getattr(self.ui, "layers_table", None)
        if isinstance(table, QTableView):
            self._table_view = table

        cells_widget = getattr(self.ui, "lstCelulas", None)
        if not isinstance(cells_widget, QListWidget):
            cells_widget = getattr(self.ui, "cells_list", None)
        if isinstance(cells_widget, QListWidget):
            self._cells_widget = cells_widget

    def set_header_view(self, header: Optional[QHeaderView]) -> None:
        self._header_view = header if isinstance(header, WordWrapHeader) else None
        if isinstance(self._header_view, WordWrapHeader):
            self._header_view.set_checkbox_section(None)
        self._refresh_selection_header()

    def _refresh_selection_header(self, *args) -> None:
        if not isinstance(self._header_view, WordWrapHeader):
            return
        section = self._header_view.checkbox_section()
        if section is None:
            return
        self._header_view.updateSection(section)

    def _connect_signals(self) -> None:
        cells_widget = self._cells_widget
        if isinstance(cells_widget, QListWidget):
            cells_widget.currentItemChanged.connect(self._on_cell_item_changed)

        name_combo = getattr(self.ui, "laminate_name_combo", None)
        if isinstance(name_combo, QComboBox):
            name_combo.currentTextChanged.connect(self._on_laminate_selected)

        if isinstance(getattr(self, "_table_view", None), QTableView):
            self._table_view.clicked.connect(self._on_table_clicked)

    def _on_cell_item_changed(
        self,
        current: Optional[QListWidgetItem],
        previous: Optional[QListWidgetItem],  # noqa: ARG002
    ) -> None:
        if current is None:
            return
        cell_id = current.data(Qt.UserRole)
        if not cell_id:
            text = current.text().split("|")[0].strip()
            cell_id = text
        self._on_cell_selected(str(cell_id))

    def _on_cell_selected(self, cell_id: str) -> None:
        if not cell_id:
            return
        self._current_cell_id = cell_id
        mapped = self.model.cell_to_laminate.get(cell_id)
        if mapped and mapped in self.model.laminados:
            self._apply_laminate(mapped)
            return
        candidates = self._laminates_by_cell.get(cell_id, [])
        if candidates:
            self._apply_laminate(candidates[0])
            return
        if self.model.laminados:
            first_name = next(iter(self.model.laminados))
            self._apply_laminate(first_name)

    def _on_laminate_selected(self, laminate_name: str) -> None:
        if self._updating:
            return
        if not laminate_name or laminate_name not in self.model.laminados:
            return
        self._apply_laminate(laminate_name)
        self._update_cell_mapping(laminate_name)

    def _on_table_clicked(self, index: QModelIndex) -> None:
        if not index.isValid():
            return
        column = index.column()
        if column == StackingTableModel.COL_SELECT:
            self.stacking_model.toggle_check(index.row())
        elif column == StackingTableModel.COL_PLY_TYPE and isinstance(self._table_view, QTableView):
            self._table_view.edit(index)
        elif column == StackingTableModel.COL_ORIENTATION and isinstance(self._table_view, QTableView):
            self._table_view.edit(index)

    def _apply_laminate(self, laminate_name: str) -> None:
        if self._updating:
            logger.debug("_apply_laminate called but _updating is True, skipping")
            return
        laminado = self.model.laminados.get(laminate_name)
        if laminado is None:
            logger.debug("_apply_laminate: laminado '%s' not found", laminate_name)
            return

        logger.debug("_apply_laminate: applying laminate '%s' (nome='%s')", laminate_name, laminado.nome)
        self._updating = True
        try:
            name_combo = getattr(self.ui, "laminate_name_combo", None)
            if isinstance(name_combo, QComboBox):
                name_combo.blockSignals(True)
                idx = name_combo.findText(laminado.nome)
                logger.debug("_apply_laminate: findText('%s') returned idx=%d", laminado.nome, idx)
                if idx >= 0:
                    name_combo.setCurrentIndex(idx)
                    logger.debug("_apply_laminate: set currentIndex to %d", idx)
                else:
                    # Name not found, add it
                    name_combo.addItem(laminado.nome)
                    name_combo.setCurrentIndex(name_combo.count() - 1)
                    logger.debug("_apply_laminate: added '%s' and set to index %d", laminado.nome, name_combo.count() - 1)
                name_combo.blockSignals(False)
                logger.debug("_apply_laminate: name_combo currentText is now '%s'", name_combo.currentText())

            color_combo = getattr(self.ui, "laminate_color_combo", None)
            if isinstance(color_combo, QComboBox):
                color_combo.blockSignals(True)
                target = str(laminado.color_index or DEFAULT_COLOR_INDEX)
                idx = color_combo.findText(target)
                if idx >= 0:
                    color_combo.setCurrentIndex(idx)
                else:
                    color_combo.addItem(target)
                    color_combo.setCurrentIndex(color_combo.count() - 1)
                color_combo.blockSignals(False)

            type_combo = getattr(self.ui, "laminate_type_combo", None)
            if isinstance(type_combo, QComboBox):
                type_combo.blockSignals(True)
                type_combo.setEditText(laminado.tipo)
                type_combo.blockSignals(False)

            tag_edit = getattr(self.ui, "laminate_tag_edit", None)
            if isinstance(tag_edit, QLineEdit):
                tag_edit.blockSignals(True)
                tag_edit.setText(getattr(laminado, "tag", ""))
                tag_edit.blockSignals(False)

            self._update_associated_cells_widget(
                self._cells_for_laminate(laminado.nome)
            )

            self.stacking_model.update_layers(laminado.camadas)
            self._current_laminate = laminado.nome
        finally:
            self._updating = False
            self._update_layers_count()
            callback = getattr(self.ui, "_on_binding_laminate_changed", None)
            if callable(callback):
                try:
                    callback(laminado.nome)
                except Exception:  # pragma: no cover - defensive
                    logger.debug(
                        "Falha ao notificar troca de laminado ativo.", exc_info=True
                    )

    def _cells_for_laminate(self, laminate_name: str) -> list[str]:
        mapped = [
            cell
            for cell in self.model.celulas_ordenadas
            if self.model.cell_to_laminate.get(cell) == laminate_name
        ]
        if mapped:
            return mapped
        laminado = self.model.laminados.get(laminate_name)
        if laminado is not None:
            return list(laminado.celulas)
        return []

    def _update_cell_mapping(self, laminate_name: str) -> None:
        cell_id = self._current_cell_id
        if not cell_id:
            return
        old = self.model.cell_to_laminate.get(cell_id)
        if old == laminate_name:
            return
        if old:
            old_lam = self.model.laminados.get(old)
            if old_lam and cell_id in old_lam.celulas:
                old_lam.celulas.remove(cell_id)
        new_lam = self.model.laminados.get(laminate_name)
        if new_lam is not None and cell_id not in new_lam.celulas:
            new_lam.celulas.append(cell_id)
        if laminate_name:
            self.model.cell_to_laminate[cell_id] = laminate_name
        else:
            self.model.cell_to_laminate.pop(cell_id, None)
        if old and old != laminate_name:
            self._update_associated_cells_text(old)
        self._update_associated_cells_text(laminate_name)
        self._laminates_by_cell = self._build_cell_index()
        if hasattr(self.ui, "_mark_dirty"):
            self.ui._mark_dirty()
        self._refresh_cell_item_label(cell_id)

    def _update_associated_cells_text(self, laminate_name: str) -> None:
        lam = self.model.laminados.get(laminate_name)
        if lam is None:
            return
        cells = self._cells_for_laminate(laminate_name)
        lam.celulas = cells
        if self._current_laminate == laminate_name:
            self._update_associated_cells_widget(cells)

    def _update_associated_cells_widget(self, cells: Sequence[str]) -> None:
        updater = getattr(self.ui, "update_associated_cells_display", None)
        if callable(updater):
            try:
                updater(list(cells))
                return
            except Exception:  # pragma: no cover - defensive
                logger.debug(
                    "Falha ao atualizar botao de celulas associadas.", exc_info=True
                )
        widget = getattr(self.ui, "associated_cells", None)
        if hasattr(widget, "setPlainText"):
            widget.setPlainText(", ".join(cells))

    def material_options(self) -> list[str]:
        materials: list[str] = []
        seen: set[str] = set()
        for laminado in self.model.laminados.values():
            for camada in laminado.camadas:
                if camada.material and camada.material not in seen:
                    seen.add(camada.material)
                    materials.append(camada.material)
        if not materials:
            materials.append("")
        return materials

    def orientation_options(self) -> list[str]:
        orientations = sorted(
            {
                float(camada.orientacao)
                for laminado in self.model.laminados.values()
                for camada in laminado.camadas
                if camada.orientacao is not None
            }
        )
        if not orientations:
            return []
        return [format_orientation_value(value) for value in orientations]

    def _most_used_material(self) -> Optional[str]:
        counts: Counter[str] = Counter()
        for laminado in self.model.laminados.values():
            for camada in laminado.camadas:
                if camada.orientacao is None:
                    continue
                material = (camada.material or "").strip()
                if material:
                    counts[material] += 1
        if not counts:
            return None
        most_common = counts.most_common(1)
        return most_common[0][0] if most_common else None

    def _refresh_cell_item_label(self, cell_id: str) -> None:
        if self._cells_widget is None:
            return
        for idx in range(self._cells_widget.count()):
            item = self._cells_widget.item(idx)
            if item.data(Qt.UserRole) == cell_id:
                self._cells_widget.blockSignals(True)
                item.setText(_format_cell_label(self.model, cell_id))
                self._cells_widget.blockSignals(False)
                break

    def _update_layers_count(self, *args) -> None:
        label = getattr(self.ui, "layers_count_label", None)
        if label is not None and hasattr(label, "setText"):
            oriented_layers = count_oriented_layers(self.stacking_model.layers())
            label.setText(
                f"Quantidade Total de Camadas: {oriented_layers}"
            )
        self._update_balance_warning()

    def _on_layers_modified(self, _layers: list[Camada]) -> None:
        if self._current_laminate and self._current_laminate in self.model.laminados:
            laminado = self.model.laminados[self._current_laminate]
            laminado.camadas = self.stacking_model.layers()
        if hasattr(self.ui, "_mark_dirty"):
            self.ui._mark_dirty()
        self._update_layers_count()
        callback = getattr(self.ui, "_on_binding_layers_modified", None)
        if callable(callback):
            try:
                callback(self._current_laminate)
            except Exception:  # pragma: no cover - defensive
                logger.debug(
                    "Falha ao notificar modificacao de camadas.", exc_info=True
                )

    def _orientation_token(self, value: object) -> Optional[float]:
        if value is None:
            return None
        try:
            return normalize_angle(value)
        except Exception:
            try:
                return normalize_angle(str(value))
            except Exception:
                return None

    def _orientations_match(self, left: Optional[float], right: Optional[float]) -> bool:
        if left is None and right is None:
            return True
        if left is None or right is None:
            return False
        return math.isclose(left, right, abs_tol=1e-6)

    def _structural_rows(self, layers: list[Camada]) -> list[int]:
        return [
            idx
            for idx, camada in enumerate(layers)
            if getattr(camada, "orientacao", None) is not None
            and is_structural_ply_label(getattr(camada, "ply_type", DEFAULT_PLY_TYPE))
        ]

    def _detect_symmetry_centers(self, layers: list[Camada]) -> tuple[bool, list[int]]:
        structural_rows = self._structural_rows(layers)
        count_struct = len(structural_rows)
        if count_struct == 0:
            return False, []
        if count_struct == 1:
            return True, [structural_rows[0]]

        i, j = 0, count_struct - 1
        while i < j:
            r_top = structural_rows[i]
            r_bot = structural_rows[j]
            camada_top = layers[r_top]
            camada_bot = layers[r_bot]
            mat_top = (getattr(camada_top, "material", "") or "").strip().lower()
            mat_bot = (getattr(camada_bot, "material", "") or "").strip().lower()
            ori_top = self._orientation_token(getattr(camada_top, "orientacao", None))
            ori_bot = self._orientation_token(getattr(camada_bot, "orientacao", None))
            if not (mat_top == mat_bot and self._orientations_match(ori_top, ori_bot)):
                return False, []
            i += 1
            j -= 1

        if count_struct % 2 == 1:
            centers = [structural_rows[count_struct // 2]]
        else:
            centers = [
                structural_rows[count_struct // 2 - 1],
                structural_rows[count_struct // 2],
            ]
        return True, centers

    def _is_unbalanced(self, layers: list[Camada]) -> bool:
        symmetric, centers = self._detect_symmetry_centers(layers)
        if not symmetric or not centers:
            return False
        structural_rows = self._structural_rows(layers)
        center_min = min(centers)
        center_set = set(centers)
        pos45 = 0
        neg45 = 0
        for row in structural_rows:
            if row in center_set:
                continue
            if row > center_min:
                continue
            orientation = self._orientation_token(getattr(layers[row], "orientacao", None))
            if orientation is None:
                continue
            if math.isclose(orientation, 45.0, abs_tol=1e-6):
                pos45 += 1
            elif math.isclose(orientation, -45.0, abs_tol=1e-6):
                neg45 += 1
        return pos45 != neg45

    def _update_balance_warning(self) -> None:
        if not hasattr(self, "stacking_model"):
            return
        layers = self.stacking_model.layers()
        warning = self._is_unbalanced(layers)
        try:
            self.stacking_model.set_unbalanced_warning(warning)
        except Exception:
            pass

    def add_layer(self, after_row: Optional[int] = None) -> bool:
        if not self._current_laminate:
            return False
        laminado = self.model.laminados.get(self._current_laminate)
        if laminado is None:
            return False
        insert_position = len(self.stacking_model.layers())
        if after_row is not None and 0 <= after_row < insert_position:
            insert_position = after_row + 1

        new_layer = Camada(
            idx=0,
            material="",
            orientacao=0,
            ativo=True,
            simetria=False,
            ply_type=DEFAULT_PLY_TYPE,
        )
        self.stacking_model.insert_layer(insert_position, new_layer)
        laminado.camadas = self.stacking_model.layers()
        self.stacking_model.clear_checks()
        self._update_layers_count()
        if hasattr(self.ui, "_mark_dirty"):
            self.ui._mark_dirty()
        return True

    def delete_checked_layers(self) -> int:
        if not self._current_laminate:
            return 0
        laminado = self.model.laminados.get(self._current_laminate)
        if laminado is None:
            return 0
        rows = self.stacking_model.checked_rows()
        if not rows:
            return 0
        removed = self.stacking_model.remove_rows(rows)
        if removed:
            laminado.camadas = self.stacking_model.layers()
            self.stacking_model.clear_checks()
            self._update_layers_count()
            if hasattr(self.ui, "_mark_dirty"):
                self.ui._mark_dirty()
        return removed

    def checked_rows(self) -> list[int]:
        return self.stacking_model.checked_rows()

    def move_selected_layer(self, direction: int) -> tuple[bool, str]:
        rows = self.checked_rows()
        if not rows:
            return False, "none"
        if len(rows) > 1:
            return False, "multi"
        current = rows[0]
        target = current + direction
        if not (0 <= target < self.stacking_model.rowCount()):
            return False, "edge"
        if not self.stacking_model.move_row(current, target):
            return False, "noop"
        laminado = self.model.laminados.get(self._current_laminate)
        if laminado is not None:
            laminado.camadas = self.stacking_model.layers()
        self._update_layers_count()
        if hasattr(self.ui, "_mark_dirty"):
            self.ui._mark_dirty()
        return True, ""


class _LayerFieldEditCommand(QUndoCommand):
    def __init__(
        self,
        model: "StackingTableModel",
        row: int,
        column: int,
        old_value: object,
        new_value: object,
        description: str,
    ) -> None:
        super().__init__(description)
        self._model = model
        self._row = row
        self._column = column
        self._old_value = old_value
        self._new_value = new_value

    def redo(self) -> None:
        self._model.apply_field_value(self._row, self._column, self._new_value)

    def undo(self) -> None:
        self._model.apply_field_value(self._row, self._column, self._old_value)


def _open_with_xlrd(file_path: Path, cause: Exception) -> _WorkbookProtocol:
    """Fallback que usa xlrd 1.2.x para planilhas .xls renomeadas."""
    if xlrd is None:
        raise ValueError(
            "A planilha parece utilizar o formato legado (.xls renomeado). "
            "Instale a dependAancia 'xlrd==1.2.0' para habilitar o fallback."
        ) from cause
    logger.warning(
        "Utilizando fallback xlrd para abrir '%s' devido a erro: %s",
        file_path,
        cause,
    )
    return _XlrdWorkbook(file_path)


class _XlrdWorkbook:
    """Wrapper simples para oferecer interface parecida com pandas.ExcelFile."""

    def __init__(self, file_path: Path) -> None:
        try:
            self._book = xlrd.open_workbook(file_path)  # type: ignore[call-arg]
        except Exception as exc:
            raise ValueError(f"NAo foi possAvel abrir '{file_path}': {exc}") from exc
        self.sheet_names = list(self._book.sheet_names())

    def parse(self, sheet_name: str) -> pd.DataFrame:
        try:
            sheet = self._book.sheet_by_name(sheet_name)
        except Exception as exc:
            raise ValueError(f"Aba '{sheet_name}' nAo pA de ser lida: {exc}") from exc

        rows: list[list[object]] = []
        for row_idx in range(sheet.nrows):
            row: list[object] = []
            for col_idx in range(sheet.ncols):
                value = sheet.cell_value(row_idx, col_idx)
                row.append(value)
            rows.append(row)

        while rows and all(_is_blank(cell) for cell in rows[0]):
            rows.pop(0)

        if not rows:
            return pd.DataFrame()

        max_len = max(len(row) for row in rows)
        normalized_rows = [
            row + [None] * (max_len - len(row)) if len(row) < max_len else row
            for row in rows
        ]
        return pd.DataFrame(normalized_rows)


def _is_blank(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    if isinstance(value, float):
        return math.isnan(value)
    return False


def _resolve_int(value: object, default: Optional[int] = None) -> Optional[int]:
    if _is_blank(value):
        return default
    try:
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return int(str(value).strip())
    except (TypeError, ValueError):
        logger.warning("Valor de Indice invalido '%s'; usando padrao.", value)
        return default


def _string_or_empty(value: object) -> str:
    if _is_blank(value):
        return ""
    return str(value).strip()


@dataclass
class _CellsSection:
    cells: list[str]
    mapping: Dict[str, str]
    separator_idx: int


def parse_cells_from_planilha1(df: pd.DataFrame) -> list[str]:
    """
    Retorna a lista de celulas listadas entre a linha 'Cells' e a linha separadora '#'.
    """
    sanitized = df.dropna(how="all").reset_index(drop=True)
    section = _extract_cells_section(sanitized)
    return section.cells


