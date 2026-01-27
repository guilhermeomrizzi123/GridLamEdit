"""Excel export helpers for GridLamEdit."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

from gridlamedit.io.spreadsheet import GridModel, save_grid_spreadsheet

logger = logging.getLogger(__name__)


def _select_sheet_name(
    available: list[str],
    preferred: str,
    *,
    context: str,
    file_name: str,
) -> str:
    if preferred in available:
        return preferred
    if not available:
        raise ValueError(f"O {context} '{file_name}' nao possui abas para leitura.")
    fallback = available[0]
    logger.warning(
        "Aba '%s' nao encontrada no %s '%s'. Usando '%s'.",
        preferred,
        context,
        file_name,
        fallback,
    )
    return fallback


def export_grid_xlsx(
    model: GridModel,
    path: Path,
    template_info: Optional[dict] = None,
) -> Path:
    """
    Persist the current GridModel into an Excel workbook matching the import layout.

    Parameters
    ----------
    model:
        GridModel instance containing the state to export.
    path:
        Target file path. When no ``.xlsx`` suffix is provided, ``.xlsx`` is used.
        Apenas ``.xlsx`` sera escrito.
    template_info:
        Placeholder for future template metadata (unused for now).

    Returns
    -------
    Path
        The resolved output path written to disk.
    """
    if model is None:
        raise ValueError("Model ausente para exportacao da planilha.")

    output_path = Path(path)
    if output_path.suffix.lower() != ".xlsx":
        output_path = output_path.with_suffix(".xlsx")
    xlsx_path = output_path

    if template_info:
        logger.info("Ignorando template_info nao utilizado: keys=%s", list(template_info))

    ensure_layers_have_material(model)
    preserved_payload = getattr(model, "preserved_columns", None)
    preserved_data = None
    preserved_columns = (3, 4, 5, 6)
    preserved_sheet = "Planilha1"
    if isinstance(preserved_payload, dict):
        preserved_data = preserved_payload.get("data")
        preserved_columns = tuple(
            preserved_payload.get("columns", preserved_columns)
        )
        preserved_sheet = preserved_payload.get("sheet_name", preserved_sheet)

    try:
        save_grid_spreadsheet(str(xlsx_path), model)
        _apply_preserved_columns(
            model,
            xlsx_path,
            preserved_data=preserved_data,
            sheet_name=preserved_sheet,
            preserved_columns=preserved_columns,
        )
    except Exception as exc:
        raise ValueError(f"Falha ao exportar arquivo .xlsx: {exc}") from exc

    return output_path


__all__ = [
    "export_grid_xlsx",
    "ensure_layers_have_material",
    "capture_preserved_columns",
]


def ensure_layers_have_material(model: GridModel) -> None:
    missing: list[str] = []

    for laminado in model.laminados.values():
        for idx, camada in enumerate(laminado.camadas, start=1):
            orientation_raw = getattr(camada, "orientacao", None)
            if orientation_raw is None:
                continue
            if isinstance(orientation_raw, str):
                orientation_text = orientation_raw.strip().lower()
                if not orientation_text or orientation_text == "empty":
                    continue

            material_text = str(getattr(camada, "material", "") or "").strip()
            if not material_text:
                missing.append(f"{laminado.nome} - camada {idx}")

    if missing:
        preview = ", ".join(missing[:5])
        remaining = len(missing) - 5
        if remaining > 0:
            preview += f" (+{remaining} camada(s))"
        raise ValueError(
            "Existem camadas sem material aplicado: "
            f"{preview}. Atribua materiais antes de exportar; o CATIA nao aceita camadas sem material."
        )


def _has_preserved_data(model: GridModel) -> bool:
    payload = getattr(model, "preserved_columns", None)
    if not isinstance(payload, dict):
        return False
    data = payload.get("data")
    return isinstance(data, list) and len(data) > 0


def _ensure_original_available(model: GridModel) -> None:
    if _has_preserved_data(model):
        return
    source_path = getattr(model, "source_excel_path", None)
    if not source_path:
        raise ValueError(
            "Nao foi possivel localizar o arquivo Excel original para preservar as colunas C-F."
        )
    original = Path(source_path)
    if not original.exists():
        raise ValueError(
            f"Arquivo Excel original nao encontrado em '{original}'. "
            "Nao foi possivel preservar as colunas C-F."
        )


def capture_preserved_columns(
    source_excel: str | Path,
    *,
    sheet_name: str = "Planilha1",
    preserved_columns: tuple[int, ...] = (3, 4, 5, 6),
) -> Optional[dict[str, object]]:
    source_path = Path(source_excel)
    if not source_path.exists():
        return None

    suffix = source_path.suffix.lower()
    if suffix == ".xls":
        data = _capture_preserved_columns_from_xls(
            source_path, sheet_name, preserved_columns
        )
    else:
        data = _capture_preserved_columns_from_xlsx(
            source_path, sheet_name, preserved_columns
        )

    if not data:
        return None

    return {
        "sheet_name": sheet_name,
        "columns": list(preserved_columns),
        "data": data,
    }


def _apply_preserved_columns(
    model: GridModel,
    output_path: Path,
    *,
    preserved_data: Optional[list[list[Optional[object]]]] = None,
    sheet_name: str = "Planilha1",
    preserved_columns: tuple[int, ...] = (3, 4, 5, 6),
) -> None:
    if preserved_data is not None:
        _restore_preserved_columns_data(
            output_path,
            sheet_name=sheet_name,
            preserved_columns=preserved_columns,
            preserved_data=preserved_data,
        )
        return

    if not getattr(model, "source_excel_path", None):
        logger.warning(
            "Exportando sem preservar colunas C-F: arquivo original nao informado."
        )
        return

    original_path = Path(model.source_excel_path)
    if not original_path.exists():
        logger.warning(
            "Exportando sem preservar colunas C-F: arquivo original nao encontrado em '%s'.",
            original_path,
        )
        return

    _restore_preserved_columns(
        original_path,
        output_path,
        sheet_name=sheet_name,
        preserved_columns=preserved_columns,
    )


def _restore_preserved_columns(
    source_excel: str | Path,
    output_path: Path,
    sheet_name: str = "Planilha1",
    preserved_columns: tuple[int, ...] = (3, 4, 5, 6),
) -> None:
    """
    Copia os valores das colunas preservadas (C-F) do arquivo original para a exportacao.
    """

    output_suffix = output_path.suffix.lower()
    if output_suffix == ".xls":
        _restore_preserved_columns_xls(
            source_excel, output_path, sheet_name, preserved_columns
        )
        return

    from openpyxl import load_workbook

    original_path = Path(source_excel)

    wb_out = load_workbook(output_path)
    out_sheet_name = _select_sheet_name(
        list(wb_out.sheetnames),
        sheet_name,
        context="arquivo exportado",
        file_name=output_path.name,
    )
    ws_out = wb_out[out_sheet_name]
    max_row_out = ws_out.max_row

    original_suffix = original_path.suffix.lower()
    if original_suffix == ".xls":
        preserved_data = _read_preserved_columns_from_xls(
            original_path, sheet_name, preserved_columns, max_row_out
        )
    else:
        preserved_data = _read_preserved_columns_from_xlsx(
            original_path, sheet_name, preserved_columns, max_row_out
        )

    for row_idx, row_values in enumerate(preserved_data, start=1):
        for col_idx, value in zip(preserved_columns, row_values, strict=True):
            ws_out.cell(row=row_idx, column=col_idx, value=value)

    wb_out.save(output_path)


def _restore_preserved_columns_xls(
    source_excel: str | Path,
    output_path: Path,
    sheet_name: str,
    preserved_columns: tuple[int, ...],
) -> None:
    """Reaplica colunas preservadas em exportacoes .xls sem usar openpyxl."""
    try:
        import xlrd  # type: ignore
        import xlwt  # type: ignore
    except ImportError as exc:  # pragma: no cover - dependencias opcionais
        raise ValueError(
            "Preservar colunas C-F de '.xls' requer as dependencias 'xlrd==1.2.0' e 'xlwt'."
        ) from exc

    original_path = Path(source_excel)

    output_book = xlrd.open_workbook(output_path)  # type: ignore[arg-type]
    output_sheet_name = _select_sheet_name(
        list(output_book.sheet_names()),
        sheet_name,
        context="arquivo exportado",
        file_name=output_path.name,
    )
    sheet_out = output_book.sheet_by_name(output_sheet_name)

    max_row_out = sheet_out.nrows

    original_suffix = original_path.suffix.lower()
    if original_suffix == ".xls":
        preserved_data = _read_preserved_columns_from_xls(
            original_path, sheet_name, preserved_columns, max_row_out
        )
    else:
        preserved_data = _read_preserved_columns_from_xlsx(
            original_path, sheet_name, preserved_columns, max_row_out
        )

    data_out: list[list[Optional[object]]] = []
    max_cols = max(sheet_out.ncols, max(preserved_columns, default=sheet_out.ncols))
    for r_idx in range(max_row_out):
        row_data = [sheet_out.cell_value(r_idx, c_idx) for c_idx in range(sheet_out.ncols)]
        # Garantir que a linha tenha colunas suficientes para sobrescrever.
        if len(row_data) < max_cols:
            row_data.extend([None] * (max_cols - len(row_data)))
        data_out.append(row_data)

    for r_idx, row_values in enumerate(preserved_data):
        while len(data_out) <= r_idx:
            data_out.append([None] * max_cols)
        row = data_out[r_idx]
        if len(row) < max_cols:
            row.extend([None] * (max_cols - len(row)))
        for col_idx, value in zip(preserved_columns, row_values, strict=True):
            dest_idx = col_idx - 1
            if dest_idx >= len(row):
                row.extend([None] * (dest_idx - len(row) + 1))
            row[dest_idx] = value

    wb_new = xlwt.Workbook()
    ws_new = wb_new.add_sheet(output_sheet_name)
    for r_idx, row in enumerate(data_out):
        for c_idx, value in enumerate(row):
            ws_new.write(r_idx, c_idx, value)
    wb_new.save(str(output_path))


def _restore_preserved_columns_data(
    output_path: Path,
    *,
    sheet_name: str,
    preserved_columns: tuple[int, ...],
    preserved_data: list[list[Optional[object]]],
) -> None:
    output_suffix = output_path.suffix.lower()
    if output_suffix == ".xls":
        _restore_preserved_columns_xls_data(
            output_path,
            sheet_name=sheet_name,
            preserved_columns=preserved_columns,
            preserved_data=preserved_data,
        )
        return

    from openpyxl import load_workbook

    wb_out = load_workbook(output_path)
    out_sheet_name = _select_sheet_name(
        list(wb_out.sheetnames),
        sheet_name,
        context="arquivo exportado",
        file_name=output_path.name,
    )
    ws_out = wb_out[out_sheet_name]

    for row_idx, row_values in enumerate(preserved_data, start=1):
        for col_idx, value in zip(preserved_columns, row_values, strict=True):
            ws_out.cell(row=row_idx, column=col_idx, value=value)

    wb_out.save(output_path)


def _restore_preserved_columns_xls_data(
    output_path: Path,
    *,
    sheet_name: str,
    preserved_columns: tuple[int, ...],
    preserved_data: list[list[Optional[object]]],
) -> None:
    """Reaplica colunas preservadas em exportacoes .xls usando dados em memoria."""
    try:
        import xlrd  # type: ignore
        import xlwt  # type: ignore
    except ImportError as exc:  # pragma: no cover - dependencias opcionais
        raise ValueError(
            "Preservar colunas C-F de '.xls' requer as dependencias 'xlrd==1.2.0' e 'xlwt'."
        ) from exc

    output_book = xlrd.open_workbook(output_path)  # type: ignore[arg-type]
    output_sheet_name = _select_sheet_name(
        list(output_book.sheet_names()),
        sheet_name,
        context="arquivo exportado",
        file_name=output_path.name,
    )
    sheet_out = output_book.sheet_by_name(output_sheet_name)

    max_row_out = sheet_out.nrows
    data_out: list[list[Optional[object]]] = []
    max_cols = max(sheet_out.ncols, max(preserved_columns, default=sheet_out.ncols))
    for r_idx in range(max_row_out):
        row_data = [sheet_out.cell_value(r_idx, c_idx) for c_idx in range(sheet_out.ncols)]
        if len(row_data) < max_cols:
            row_data.extend([None] * (max_cols - len(row_data)))
        data_out.append(row_data)

    for r_idx, row_values in enumerate(preserved_data):
        while len(data_out) <= r_idx:
            data_out.append([None] * max_cols)
        row = data_out[r_idx]
        if len(row) < max_cols:
            row.extend([None] * (max_cols - len(row)))
        for col_idx, value in zip(preserved_columns, row_values, strict=True):
            dest_idx = col_idx - 1
            if dest_idx >= len(row):
                row.extend([None] * (dest_idx - len(row) + 1))
            row[dest_idx] = value

    wb_new = xlwt.Workbook()
    ws_new = wb_new.add_sheet(output_sheet_name)
    for r_idx, row in enumerate(data_out):
        for c_idx, value in enumerate(row):
            ws_new.write(r_idx, c_idx, value)
    wb_new.save(str(output_path))


def _read_preserved_columns_from_xlsx(
    original_path: Path,
    sheet_name: str,
    preserved_columns: tuple[int, ...],
    max_rows: int,
) -> list[list[Optional[object]]]:
    from openpyxl import load_workbook

    wb_in = load_workbook(original_path, data_only=False, read_only=False)
    input_sheet_name = _select_sheet_name(
        list(wb_in.sheetnames),
        sheet_name,
        context="arquivo original",
        file_name=original_path.name,
    )
    ws_in = wb_in[input_sheet_name]
    max_row = min(max_rows, ws_in.max_row)
    data: list[list[Optional[object]]] = []
    for row_idx in range(1, max_row + 1):
        row_data: list[Optional[object]] = []
        for col_idx in preserved_columns:
            row_data.append(ws_in.cell(row=row_idx, column=col_idx).value)
        data.append(row_data)
    return data


def _capture_preserved_columns_from_xlsx(
    original_path: Path,
    sheet_name: str,
    preserved_columns: tuple[int, ...],
) -> list[list[Optional[object]]]:
    from openpyxl import load_workbook

    wb_in = load_workbook(original_path, data_only=False, read_only=False)
    input_sheet_name = _select_sheet_name(
        list(wb_in.sheetnames),
        sheet_name,
        context="arquivo original",
        file_name=original_path.name,
    )
    ws_in = wb_in[input_sheet_name]
    data: list[list[Optional[object]]] = []
    for row_idx in range(1, ws_in.max_row + 1):
        row_data: list[Optional[object]] = []
        for col_idx in preserved_columns:
            row_data.append(ws_in.cell(row=row_idx, column=col_idx).value)
        data.append(row_data)
    return data


def _read_preserved_columns_from_xls(
    original_path: Path,
    sheet_name: str,
    preserved_columns: tuple[int, ...],
    max_rows: int,
) -> list[list[Optional[object]]]:
    try:
        import xlrd  # type: ignore
    except ImportError as exc:  # pragma: no cover - dependAancia opcional
        raise ValueError(
            "Preservar colunas C-F de arquivos '.xls' requer a dependAancia 'xlrd==1.2.0'."
        ) from exc

    workbook = xlrd.open_workbook(original_path)  # type: ignore[arg-type]
    input_sheet_name = _select_sheet_name(
        list(workbook.sheet_names()),
        sheet_name,
        context="arquivo original",
        file_name=original_path.name,
    )
    sheet = workbook.sheet_by_name(input_sheet_name)

    max_row = min(max_rows, sheet.nrows)
    data: list[list[Optional[object]]] = []
    for row_idx in range(max_row):
        row_data: list[Optional[object]] = []
        for col_idx in preserved_columns:
            if col_idx - 1 < sheet.ncols:
                cell_value = sheet.cell_value(row_idx, col_idx - 1)
            else:
                cell_value = None
            row_data.append(cell_value)
        data.append(row_data)
    return data


def _capture_preserved_columns_from_xls(
    original_path: Path,
    sheet_name: str,
    preserved_columns: tuple[int, ...],
) -> list[list[Optional[object]]]:
    try:
        import xlrd  # type: ignore
    except ImportError as exc:  # pragma: no cover - dependAancia opcional
        raise ValueError(
            "Preservar colunas C-F de arquivos '.xls' requer a dependAancia 'xlrd==1.2.0'."
        ) from exc

    workbook = xlrd.open_workbook(original_path)  # type: ignore[arg-type]
    input_sheet_name = _select_sheet_name(
        list(workbook.sheet_names()),
        sheet_name,
        context="arquivo original",
        file_name=original_path.name,
    )
    sheet = workbook.sheet_by_name(input_sheet_name)

    data: list[list[Optional[object]]] = []
    for row_idx in range(sheet.nrows):
        row_data: list[Optional[object]] = []
        for col_idx in preserved_columns:
            if col_idx - 1 < sheet.ncols:
                cell_value = sheet.cell_value(row_idx, col_idx - 1)
            else:
                cell_value = None
            row_data.append(cell_value)
        data.append(row_data)
    return data
