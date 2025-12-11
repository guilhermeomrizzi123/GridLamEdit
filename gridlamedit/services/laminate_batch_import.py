"""Helpers to manage the batch laminate Excel template."""

from __future__ import annotations

import math
import re
import tempfile
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

from gridlamedit.io.spreadsheet import normalize_angle


@dataclass
class BatchLaminateInput:
    """Raw laminate data captured from the batch template."""

    tag: str
    is_symmetric: bool
    center_is_single: bool
    orientations: list[Optional[float]]
    row_number: int


def _normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def _is_blank(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _is_empty_token(value: object) -> bool:
    if isinstance(value, str) and value.strip().lower() in {"x", "empty"}:
        return True
    return False


def _truthy(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return False
        return bool(value)
    token = _normalize_header(value)
    return token in {"y", "yes", "sim", "s", "true", "1"}


def create_blank_batch_template(
    source_path: Path,
    *,
    destination: Optional[Path] = None,
    sheet_name: Optional[str] = None,
) -> Path:
    """Return a cleaned copy of the batch template with data rows cleared."""

    if not source_path.exists():
        raise FileNotFoundError(f"Template nao encontrado em '{source_path}'.")

    target_path = destination or Path(tempfile.gettempdir()) / (
        f"gridlam_batch_template_{uuid.uuid4().hex}.xlsx"
    )
    target_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(source_path)
    sheet = None
    if sheet_name and sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    elif workbook.sheetnames:
        sheet = workbook[workbook.sheetnames[0]]
    if sheet is None:
        raise ValueError("Nenhuma planilha encontrada no template de lote.")

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            cell.value = None

    workbook.save(target_path)
    return target_path


def parse_batch_template(
    path: Path,
    *,
    sheet: Optional[str] = None,
) -> list[BatchLaminateInput]:
    """Read the filled template and return laminate inputs."""

    file_path = Path(path)
    if not file_path.exists():
        raise FileNotFoundError(f"Arquivo '{file_path}' nao encontrado.")

    df = pd.read_excel(file_path, sheet_name=sheet or 0)
    if df.empty:
        return []
    df = df.where(pd.notna(df), None)

    normalized = {_normalize_header(col): col for col in df.columns}
    tag_col = None
    for key, original in normalized.items():
        if key == "tag" or key.endswith("tag"):
            tag_col = original
            break

    symmetry_candidates = {"simmetry", "symmetry", "simetria"}
    symmetry_col = None
    for key, original in normalized.items():
        if key in symmetry_candidates:
            symmetry_col = original
            break

    center_candidates = {
        "lastsequenceassimmetrycenter",
        "lastsequenceasimmetrycenter",
        "lastsequenceasymmetrycenter",
        "sequenciacentral",
        "centerflag",
        "centersingle",
    }
    center_col = None
    for key, original in normalized.items():
        if key in center_candidates or ("center" in key and "sequence" in key):
            center_col = original
            break

    orientation_cols: list[tuple[int, object]] = []
    for col in df.columns:
        text = str(col).strip()
        if text.isdigit():
            orientation_cols.append((int(text), col))
    orientation_cols.sort(key=lambda item: item[0])

    if not orientation_cols:
        raise ValueError("Nenhuma coluna numerica de sequencia encontrada no template.")

    laminates: list[BatchLaminateInput] = []
    for idx, row in df.iterrows():
        row_number = idx + 2  # header is row 1
        orientations: list[Optional[float]] = []
        has_orientation = False
        for _, label in orientation_cols:
            raw = row.get(label)
            if _is_blank(raw) or _is_empty_token(raw):
                orientations.append(None)
                continue
            try:
                value = normalize_angle(raw)
            except ValueError as exc:
                raise ValueError(
                    f"Linha {row_number}: orientacao invalida em '{label}': {exc}"
                ) from exc
            orientations.append(value)
            has_orientation = True

        if not has_orientation and all(val is None for val in orientations):
            continue

        tag_value = str(row.get(tag_col) or "").strip() if tag_col else ""
        is_symmetric = _truthy(row.get(symmetry_col)) if symmetry_col else False
        center_value = row.get(center_col) if center_col else None
        laminates.append(
            BatchLaminateInput(
                tag=tag_value,
                is_symmetric=is_symmetric,
                center_is_single=_truthy(center_value),
                orientations=orientations,
                row_number=row_number,
            )
        )

    return laminates


__all__ = [
    "BatchLaminateInput",
    "create_blank_batch_template",
    "parse_batch_template",
]
